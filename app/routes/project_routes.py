"""
project_routes.py - مسارات إدارة المشاريع المتكاملة
"""

from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, session,current_app,g,send_file
from flask_login import login_required, current_user
from app.models import db
from app.routes import project_bp
from app.models import (
    Organization,Notification,Project, ProjectLocation, ProjectDates, ProjectBudget, ProjectCost,
    ProjectPerformance, ProjectProgress, ProjectStatistics, ProjectCodeDictionary,ProjectCodeValue,ProjectCodeAssignment, ProjectUDF,
    EPS, OBS, Calendar, BudgetLog, FundingSource, SpendingPlanItem,Meeting,
    WBS, Activity, Task, NotebookEntry, Milestone, Client, Consultant, Supplier,ResourceRequestUpdate,Issue,TaskPlanning,TaskExecution,
    User,ActivityResource,Resource,Risk,ProjectDocument,ResourceDelivery,ResourceRequest, ResourceRequestItem, ResourceRequestNotification
)
from app.services.notification_service import NotificationService
from app.services.resource_service import ResourceService
from app.services.smart_monitor import SmartMonitoringSystem
from app.services.resource_request_service import ResourceRequestService
from datetime import datetime, date,timedelta
import json
import uuid
import os
from werkzeug.utils import secure_filename
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.services.update_service import UpdateService
import logging

# ✅ إضافة تعريف logger
logger = logging.getLogger(__name__)

# ============================================
# دوال مساعدة
# ============================================
@project_bp.before_request
def load_company():
    if current_user.is_authenticated:
        g.company = Organization.query.get(current_user.org_id)
        g.notifications_count = Notification.query.filter_by(
            user_id=current_user.id, 
            is_read=False
        ).count()
        g.delayed_tasks_count = Task.query.join(Project).filter(
            Task.status.in_(['pending', 'in_progress']),
            TaskPlanning.planned_finish < date.today()
        ).count()
        g.pending_deliveries_count = ResourceDelivery.query.filter_by(
            status='pending'
        ).count() if current_user.role in ['org_admin', 'project_manager'] else 0
        
        # إضافة إحصائيات الموارد
        g.low_stock_resources = Resource.query.filter(
            Resource.available_quantity < Resource.minimum_quantity
        ).count() if hasattr(Resource, 'minimum_quantity') else 0
        # ⭐ إحصائيات الاجتماعات القادمة
        # ========== الاجتماعات القادمة ==========
        today = datetime.now().date()
        next_week = today + timedelta(days=7)
        
        upcoming_meetings_query = Meeting.query.filter(
            Meeting.project.has(org_id=current_user.org_id),
            Meeting.scheduled_date >= today,
            Meeting.scheduled_date <= next_week,
            Meeting.status == 'scheduled'
        ).order_by(Meeting.scheduled_date)
        
        # الحصول على القائمة للعرض
        g.upcoming_meetings = upcoming_meetings_query.limit(5).all()
        
        # ✅ حساب العدد بشكل صحيح
        g.upcoming_meetings_count = upcoming_meetings_query.count()
        
        # ========== القضايا المفتوحة ==========
        open_issues_count = Issue.query.join(Project).filter(
            Project.org_id == current_user.org_id,
            Issue.status.in_(['open', 'in_progress'])
        ).count()
        g.open_issues_count = open_issues_count
        
        # آخر 5 قضايا
        recent_issues = Issue.query.join(Project).filter(
            Project.org_id == current_user.org_id
        ).order_by(Issue.reported_date.desc()).limit(5).all()
        g.recent_issues = recent_issues
        
        # إحصائيات القضايا
        g.issues_stats = {
            'open': Issue.query.join(Project).filter(
                Project.org_id == current_user.org_id, 
                Issue.status == 'open'
            ).count(),
            'in_progress': Issue.query.join(Project).filter(
                Project.org_id == current_user.org_id, 
                Issue.status == 'in_progress'
            ).count(),
            'total': Issue.query.join(Project).filter(
                Project.org_id == current_user.org_id
            ).count()
        }
    else:
        g.company = None
        g.upcoming_meetings =None
        g.recent_issues=None
        g.delayed_tasks_count = 0
        g.notifications_count = 0
        g.pending_deliveries_count=0
        g.low_stock_resources=0
        g.upcoming_meetings_count=0
        g.open_issues_count =0
        g.issues_stats={}
def get_org_id():
    """الحصول على معرف المؤسسة"""
    return current_user.org_id

def check_project_access(project_id):
    """التحقق من صلاحية الوصول للمشروع"""
    project = Project.query.get_or_404(project_id)
    if project.created_by != current_user.id and current_user.role != 'admin':
        return None
    return project

def get_project_budget_summary(project_id):
    """الحصول على ملخص ميزانية المشروع"""
    project = Project.query.get(project_id)
    if not project:
        return {}
    
    activities = Activity.query.filter_by(project_id=project_id).all()
    distributed_budget = sum(a.planned_value or 0 for a in activities)
    
    current_budget = project.budget.current_budget if project.budget else 0
    unallocated_budget = max(0, current_budget - distributed_budget)
    
    return {
        'current_budget': current_budget,
        'unallocated_budget': unallocated_budget,
        'distributed_budget': distributed_budget,
        'actual_cost': project.cost.total_actual_cost if project.cost else 0,
        'current_variance': current_budget - (project.cost.total_actual_cost if project.cost else 0)
    }

def get_project_statistics(project_id):
    """تحديث وجلب إحصائيات المشروع"""
    project = Project.query.get(project_id)
    if not project:
        return {}
    
    if project.statistics:
        project.statistics.update()
        db.session.commit()
        return project.statistics
    
    return {}
# أضف هذه الدوال المساعدة في بداية الملف

def update_project_statistics(project_id):
    """تحديث إحصائيات المشروع"""
    project = Project.query.get(project_id)
    if not project:
        return
    
    if project.statistics:
        project.statistics.total_activities = Activity.query.filter_by(project_id=project_id).count()
        project.statistics.completed_activities = Activity.query.filter_by(project_id=project_id, status='completed').count()
        project.statistics.in_progress_activities = Activity.query.filter_by(project_id=project_id, status='in_progress').count()
        project.statistics.not_started_activities = Activity.query.filter_by(project_id=project_id, status='not_started').count()
        project.statistics.critical_activities = Activity.query.filter_by(project_id=project_id, is_critical=True).count()
        
        project.statistics.total_tasks = Task.query.filter_by(project_id=project_id).count()
        project.statistics.completed_tasks = Task.query.filter_by(project_id=project_id, status='completed').count()
        
        project.statistics.last_calculated = datetime.utcnow()
        db.session.commit()


def calculate_spending_totals(project_id):
    """حساب إجماليات خطة الصرف"""
    items = SpendingPlanItem.query.filter_by(project_id=project_id).order_by(SpendingPlanItem.date).all()
    
    running_spending = 0
    running_benefit = 0
    
    for item in items:
        running_spending += item.planned_amount
        running_benefit += item.benefit_amount
        item.spending_tally = running_spending
        item.benefit_tally = running_benefit
        item.undistributed_variance = running_spending - running_benefit
        item.benefit_variance = item.benefit_amount - item.planned_amount
    
    db.session.commit()
    
    return {
        'total_spending': running_spending,
        'total_benefit': running_benefit,
        'undistributed_variance': running_spending - running_benefit,
        'benefit_variance': running_benefit - running_spending
    }


def get_spending_plan_data(project_id):
    """الحصول على بيانات خطة الصرف للتقرير"""
    items = SpendingPlanItem.query.filter_by(project_id=project_id).order_by(SpendingPlanItem.date).all()
    
    result = []
    for item in items:
        result.append({
            'date': item.date.strftime('%Y-%m-%d'),
            'planned_amount': item.planned_amount,
            'benefit_amount': item.benefit_amount,
            'spending_tally': item.spending_tally if hasattr(item, 'spending_tally') else 0,
            'benefit_tally': item.benefit_tally if hasattr(item, 'benefit_tally') else 0,
            'undistributed_variance': item.undistributed_variance if hasattr(item, 'undistributed_variance') else 0,
            'benefit_variance': item.benefit_variance if hasattr(item, 'benefit_variance') else 0
        })
    
    return result

def get_project_resource_stats(project_id):
    """الحصول على إحصائيات موارد المشروع"""
    assignments = ActivityResource.query.join(Activity).filter(Activity.project_id == project_id).all()
    
    stats = {
        'labor_count': 0,
        'equipment_count': 0,
        'material_count': 0,
        'total_cost': 0,
        'total_units': 0
    }
    
    for ass in assignments:
        if ass.resource:
            if ass.resource.resource_type == 'labor':
                stats['labor_count'] += 1
            elif ass.resource.resource_type == 'equipment':
                stats['equipment_count'] += 1
            elif ass.resource.resource_type == 'material':
                stats['material_count'] += 1
            
            stats['total_cost'] += ass.actual_cost or 0
            stats['total_units'] += ass.actual_quantity or 0
    
    return stats


def get_resource_chart_data(project_id):
    """الحصول على بيانات الرسم البياني للموارد"""
    project = Project.query.get(project_id)
    if not project:
        return []
    
    resources = Resource.query.filter_by(org_id=project.org_id).all()
    chart_data = []
    
    for r in resources[:10]:  # Limit to 10 resources for chart
        assigned = sum(a.planned_quantity or 0 for a in ActivityResource.query.filter_by(resource_id=r.id)
                      .join(Activity).filter(Activity.project_id == project_id).all())
        chart_data.append({
            'name': r.name[:15] + '...' if len(r.name) > 15 else r.name,
            'available': r.available_quantity or 0,
            'assigned': assigned
        })
    
    return chart_data
# ============================================
# صفحات المشاريع
# ============================================

@project_bp.route('/')
@login_required
def list_projects():
    """عرض قائمة المشاريع"""
    projects = Project.query.filter_by(created_by=current_user.id).order_by(Project.created_at.desc()).all()
    
    # إحصائيات سريعة
    stats = {
        'total': len(projects),
        'active': sum(1 for p in projects if p.status == 'active'),
        'planning': sum(1 for p in projects if p.status == 'planning'),
        'completed': sum(1 for p in projects if p.status == 'completed'),
        'overdue': sum(1 for p in projects if p.is_overdue)
    }
    
    return render_template('projects/list.html', projects=projects, stats=stats)


@project_bp.route('/<int:project_id>')
@login_required
def project_detail(project_id):
    """عرض تفاصيل المشروع مع جميع التبويبات"""
    project = check_project_access(project_id)
    if not project:
        flash('غير مصرح بالوصول', 'danger')
        return redirect(url_for('projects.list_projects'))
    
    # تحديث الإحصائيات
    get_project_statistics(project_id)
    
    # جلب البيانات المرتبطة
    eps_list = EPS.query.filter_by(org_id=get_org_id()).all()
    obs_list = OBS.query.filter_by(org_id=get_org_id()).all()
    calendars = Calendar.query.filter_by(org_id=get_org_id()).all()
    users = User.query.filter_by(org_id=get_org_id()).all()
    clients = User.query.filter_by(org_id=get_org_id()).all()
    consultants = User.query.filter_by(org_id=get_org_id()).all()
    suppliers = User.query.filter_by(org_id=get_org_id()).all()
    
    # إحصائيات المشروع
    stats = {
        'total_activities': Activity.query.filter_by(project_id=project_id).count(),
        'total_tasks': Task.query.filter_by(project_id=project_id).count(),
        'total_milestones': Milestone.query.filter_by(project_id=project_id).count(),
        'total_funding': sum(f.amount for f in project.funding_sources),
        'notebook_entries': project.notebook_entries.count(),
        'risks': len(project.risks),
        'issues': len(project.issues)
        
    }
    
    budget_summary = get_project_budget_summary(project_id)
    
    # معالم المشروع
    milestones = Milestone.query.filter_by(project_id=project_id).order_by(Milestone.planned_date).all()
    
    # سجل الميزانية
    budget_logs = BudgetLog.query.filter_by(project_id=project_id).order_by(BudgetLog.date.desc()).limit(10).all()
    
    # مصادر التمويل
    funding_sources = FundingSource.query.filter_by(project_id=project_id).all()
    
    # خطة الصرف
    # بيانات خطة الصرف
    spending_plan = SpendingPlanItem.query.filter_by(project_id=project_id).order_by(SpendingPlanItem.date).all()
    
    # حساب إجماليات خطة الصرف
    total_spending_plan = sum(i.planned_amount for i in spending_plan)
    total_benefit_plan = sum(i.benefit_amount for i in spending_plan)
    
    # حساب الفروق
    if spending_plan:
        last_item = spending_plan[-1]
        undistributed_variance = last_item.undistributed_variance if hasattr(last_item, 'undistributed_variance') else total_spending_plan - total_benefit_plan
        benefit_variance = total_benefit_plan - total_spending_plan
    else:
        undistributed_variance = 0
        benefit_variance = 0
    
    # أكواد المشروع
    # الحصول على جميع قواميس الأكواد للمؤسسة
    code_dictionaries = ProjectCodeDictionary.query.filter_by(
        org_id=current_user.org_id,
        is_active=True
    ).all()

    # تجهيز بيانات الأكواد مع القيم المرتبطة
    project_codes = []
    for dictionary in code_dictionaries:
        # البحث عن الكود المرتبط بهذا المشروع لهذا القاموس
        assignment = ProjectCodeAssignment.query.filter_by(
            project_id=project_id,
            dictionary_id=dictionary.id
        ).first()
        
        project_codes.append({
            'dictionary': dictionary,
            'assigned_code': assignment.code_value if assignment else None,
            'assigned_code_id': assignment.code_value_id if assignment else None
        })
    
    # الحقول المخصصة
    project_udfs = ProjectUDF.query.filter_by(project_id=project_id).all()
    # جلب جميع الأنشطة في المشروع
    all_resources=Resource.query.filter_by(org_id=project.org_id).all()
    activities=Activity.query.filter_by(project_id=project_id).all()
    # Get resource assignments
    # جلب التخصيصات الحالية
    resource_assignments = []
    resource_stats = {
        'labor_count': 0,
        'equipment_count': 0,
        'material_count': 0,
        'labor_units': 0,
        'equipment_units': 0,
        'material_units': 0,
        'total_planned_cost': 0,
        'total_actual_cost': 0
    }
    
    for activity in activities:
        for assignment in activity.resources:
            resource = assignment.resource
            planned_cost = assignment.planned_quantity * resource.cost_per_unit
            actual_cost = assignment.actual_quantity * resource.cost_per_unit
            
            resource_assignments.append({
                'id': assignment.id,
                'resource_id': resource.id,
                'resource_code': resource.resource_id,
                'resource_name': resource.name,
                'resource_type': resource.resource_type,
                'activity_id': activity.id,
                'activity_code': activity.activity_id,
                'activity_name': activity.activity_name,
                'planned_units': assignment.planned_quantity,
                'actual_units': assignment.actual_quantity,
                'remaining_units': assignment.remaining_quantity,
                'planned_cost': planned_cost,
                'actual_cost': actual_cost,
                'rate_type': getattr(assignment, 'rate_type', 'Standard'),
                'drive_dates': getattr(assignment, 'drive_dates', True)
            })
            
            # تحديث الإحصائيات
            if resource.resource_type == 'labor':
                resource_stats['labor_count'] += 1
                resource_stats['labor_units'] += assignment.planned_quantity
            elif resource.resource_type == 'equipment':
                resource_stats['equipment_count'] += 1
                resource_stats['equipment_units'] += assignment.planned_quantity
            elif resource.resource_type == 'material':
                resource_stats['material_count'] += 1
                resource_stats['material_units'] += assignment.planned_quantity
            
            resource_stats['total_planned_cost'] += planned_cost
            resource_stats['total_actual_cost'] += actual_cost
    
    # بيانات الرسم البياني للموارد
    resource_chart_data =get_resource_chart_data(project_id)
    
    # إعدادات المشروع الافتراضية للموارد
    project_defaults = {
        'rate_type': getattr(project, 'resource_defaults', {}).get('rate_type', 'Standard Rate'),
        'drive_activity_dates': getattr(project, 'resource_defaults', {}).get('drive_activity_dates', True),
        'allow_multiple_assignments': getattr(project, 'resource_defaults', {}).get('allow_multiple_assignments', True)
    }

    return render_template('projects/detail11.html',
                         project=project,
                         eps_list=eps_list,
                         obs_list=obs_list,
                         calendars=calendars,
                         users=users,
                         clients=clients,
                         consultants=consultants,
                         suppliers=suppliers,
                         stats=stats,
                         budget_summary=budget_summary,
                         milestones=milestones,
                         budget_logs=budget_logs,
                         funding_sources=funding_sources,
                         spending_plan=spending_plan,
                         total_spending_plan=total_spending_plan,
                         total_benefit_plan=total_benefit_plan,
                         undistributed_variance=undistributed_variance,
                         benefit_variance=benefit_variance,
                         project_codes=project_codes,
                         project_udfs=project_udfs,
                         resource_assignments=resource_assignments,
                         resource_stats=resource_stats,
                         resource_chart_data=resource_chart_data,
                         all_resources=all_resources,
                         activities=activities,
                         project_defaults=project_defaults,now=datetime.now)

@project_bp.route('/apis/<int:project_id>')
@login_required
def project_detail22(project_id):
    """عرض تفاصيل المشروع مع جميع التبويبات"""
    project = check_project_access(project_id)
    if not project:
        flash('غير مصرح بالوصول', 'danger')
        return redirect(url_for('projects.list_projects'))
    
    # تحديث الإحصائيات
    get_project_statistics(project_id)
    
    # جلب البيانات المرتبطة
    eps_list = EPS.query.filter_by(org_id=get_org_id()).all()
    obs_list = OBS.query.filter_by(org_id=get_org_id()).all()
    calendars = Calendar.query.filter_by(org_id=get_org_id()).all()
    users = User.query.filter_by(org_id=get_org_id()).all()
    clients = User.query.filter_by(org_id=get_org_id()).all()
    consultants = User.query.filter_by(org_id=get_org_id()).all()
    suppliers = User.query.filter_by(org_id=get_org_id()).all()
    
    # إحصائيات المشروع
    stats = {
        'total_activities': Activity.query.filter_by(project_id=project_id).count(),
        'total_tasks': Task.query.filter_by(project_id=project_id).count(),
        'total_milestones': Milestone.query.filter_by(project_id=project_id).count(),
        'total_funding': sum(f.amount for f in project.funding_sources),
        'notebook_entries': project.notebook_entries.count(),
        'risks': len(project.risks),
        'issues': len(project.issues)
        
    }
    
    budget_summary = get_project_budget_summary(project_id)
    
    # معالم المشروع
    milestones = Milestone.query.filter_by(project_id=project_id).order_by(Milestone.planned_date).all()
    
    # سجل الميزانية
    budget_logs = BudgetLog.query.filter_by(project_id=project_id).order_by(BudgetLog.date.desc()).limit(10).all()
    
    # مصادر التمويل
    funding_sources = FundingSource.query.filter_by(project_id=project_id).all()
    
    # خطة الصرف
    # بيانات خطة الصرف
    spending_plan = SpendingPlanItem.query.filter_by(project_id=project_id).order_by(SpendingPlanItem.date).all()
    
    # حساب إجماليات خطة الصرف
    total_spending_plan = sum(i.planned_amount for i in spending_plan)
    total_benefit_plan = sum(i.benefit_amount for i in spending_plan)
    
    # حساب الفروق
    if spending_plan:
        last_item = spending_plan[-1]
        undistributed_variance = last_item.undistributed_variance if hasattr(last_item, 'undistributed_variance') else total_spending_plan - total_benefit_plan
        benefit_variance = total_benefit_plan - total_spending_plan
    else:
        undistributed_variance = 0
        benefit_variance = 0
    
    # أكواد المشروع
    # الحصول على جميع قواميس الأكواد للمؤسسة
    code_dictionaries = ProjectCodeDictionary.query.filter_by(
        org_id=current_user.org_id,
        is_active=True
    ).all()

    # تجهيز بيانات الأكواد مع القيم المرتبطة
    project_codes = []
    for dictionary in code_dictionaries:
        # البحث عن الكود المرتبط بهذا المشروع لهذا القاموس
        assignment = ProjectCodeAssignment.query.filter_by(
            project_id=project_id,
            dictionary_id=dictionary.id
        ).first()
        
        project_codes.append({
            'dictionary': dictionary,
            'assigned_code': assignment.code_value if assignment else None,
            'assigned_code_id': assignment.code_value_id if assignment else None
        })
    
    # الحقول المخصصة
    project_udfs = ProjectUDF.query.filter_by(project_id=project_id).all()
    # جلب جميع الأنشطة في المشروع
    all_resources=Resource.query.filter_by(org_id=project.org_id).all()
    activities=Activity.query.filter_by(project_id=project_id).all()
    # Get resource assignments
    # جلب التخصيصات الحالية
    resource_assignments = []
    resource_stats = {
        'labor_count': 0,
        'equipment_count': 0,
        'material_count': 0,
        'labor_units': 0,
        'equipment_units': 0,
        'material_units': 0,
        'total_planned_cost': 0,
        'total_actual_cost': 0
    }
    
    for activity in activities:
        for assignment in activity.resources:
            resource = assignment.resource
            planned_cost = assignment.planned_quantity * resource.cost_per_unit
            actual_cost = assignment.actual_quantity * resource.cost_per_unit
            
            resource_assignments.append({
                'id': assignment.id,
                'resource_id': resource.id,
                'resource_code': resource.resource_id,
                'resource_name': resource.name,
                'resource_type': resource.resource_type,
                'activity_id': activity.id,
                'activity_code': activity.activity_id,
                'activity_name': activity.activity_name,
                'planned_units': assignment.planned_quantity,
                'actual_units': assignment.actual_quantity,
                'remaining_units': assignment.remaining_quantity,
                'planned_cost': planned_cost,
                'actual_cost': actual_cost,
                'rate_type': getattr(assignment, 'rate_type', 'Standard'),
                'drive_dates': getattr(assignment, 'drive_dates', True)
            })
            
            # تحديث الإحصائيات
            if resource.resource_type == 'labor':
                resource_stats['labor_count'] += 1
                resource_stats['labor_units'] += assignment.planned_quantity
            elif resource.resource_type == 'equipment':
                resource_stats['equipment_count'] += 1
                resource_stats['equipment_units'] += assignment.planned_quantity
            elif resource.resource_type == 'material':
                resource_stats['material_count'] += 1
                resource_stats['material_units'] += assignment.planned_quantity
            
            resource_stats['total_planned_cost'] += planned_cost
            resource_stats['total_actual_cost'] += actual_cost
    
    # بيانات الرسم البياني للموارد
    resource_chart_data =get_resource_chart_data(project_id)
    
    # إعدادات المشروع الافتراضية للموارد
    project_defaults = {
        'rate_type': getattr(project, 'resource_defaults', {}).get('rate_type', 'Standard Rate'),
        'drive_activity_dates': getattr(project, 'resource_defaults', {}).get('drive_activity_dates', True),
        'allow_multiple_assignments': getattr(project, 'resource_defaults', {}).get('allow_multiple_assignments', True)
    }

    return render_template('projects/detail22.html',
                         project=project,
                         eps_list=eps_list,
                         obs_list=obs_list,
                         calendars=calendars,
                         users=users,
                         clients=clients,
                         consultants=consultants,
                         suppliers=suppliers,
                         stats=stats,
                         budget_summary=budget_summary,
                         milestones=milestones,
                         budget_logs=budget_logs,
                         funding_sources=funding_sources,
                         spending_plan=spending_plan,
                         total_spending_plan=total_spending_plan,
                         total_benefit_plan=total_benefit_plan,
                         undistributed_variance=undistributed_variance,
                         benefit_variance=benefit_variance,
                         project_codes=project_codes,
                         project_udfs=project_udfs,
                         resource_assignments=resource_assignments,
                         resource_stats=resource_stats,
                         resource_chart_data=resource_chart_data,
                         all_resources=all_resources,
                         activities=activities,
                         project_defaults=project_defaults,now=datetime.now)


@project_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_project():
    """إنشاء مشروع جديد"""
    eps_ids = request.args.get('eps_id', type=int) 
    eps_id=None
    if eps_ids:
        eps_id=EPS.query.get_or_404(eps_ids)
    else:
        eps_id=None
    if request.method == 'POST':
        try:
            # التحقق من عدم تكرار كود المشروع
            existing = Project.query.filter_by(
                project_code=request.form.get('project_code')
            ).first()
            
            if existing:
                flash('كود المشروع موجود مسبقاً', 'danger')
                return redirect(url_for('projects.create_project'))
            
            # إنشاء المشروع الرئيسي
            project = Project(
                name=request.form.get('name'),
                project_code=request.form.get('project_code'),
                description=request.form.get('description'),
                org_id=get_org_id(),
                eps_id=request.form.get('eps_id') or None,
                obs_id=request.form.get('obs_id') or None,
                calendar_id=request.form.get('calendar_id') or None,
                financial_calendar_id=request.form.get('financial_calendar_id') or None,
                status=request.form.get('status', 'planning'),
                priority_level=int(request.form.get('priority_level', 50)),
                risk_level=request.form.get('risk_level', 'medium'),
                complexity=request.form.get('complexity', 'medium'),
                category=request.form.get('category'),
                project_type=request.form.get('project_type'),
                project_scale=request.form.get('project_scale'),
                is_template=request.form.get('is_template') == 'on',
                client_id=request.form.get('client_id') or None,
                consultant_id=request.form.get('consultant_id') or None,
                project_manager_id=request.form.get('project_manager_id') or None,
                checked_out_by=request.form.get('assistant_manager_id') or None,
                website=request.form.get('website'),
                fiscal_year_start=request.form.get('fiscal_year_start', '01-01'),
                critical_definition=request.form.get('critical_definition', 'Total Float <= 0'),
                base_currency=request.form.get('base_currency'),
                created_by=current_user.id,
                uuid=str(uuid.uuid4())
            )
            db.session.add(project)
            db.session.flush()
            
            # إنشاء الموقع
            if any([request.form.get('site_name'), request.form.get('city'), 
                    request.form.get('country'), request.form.get('address')]):
                location = ProjectLocation(
                    project_id=project.id,
                    site_name=request.form.get('site_name'),
                    city=request.form.get('city'),
                    country=request.form.get('country'),
                    address=request.form.get('address'),
                    coordinates=request.form.get('coordinates')
                )
                db.session.add(location)
            
            # إنشاء التواريخ
            dates = ProjectDates(
                project_id=project.id,
                planned_start=datetime.strptime(request.form.get('planned_start'), '%Y-%m-%d') if request.form.get('planned_start') else None,
                planned_finish=datetime.strptime(request.form.get('planned_finish'), '%Y-%m-%d') if request.form.get('planned_finish') else None,
                must_finish_by=datetime.strptime(request.form.get('must_finish_by'), '%Y-%m-%d') if request.form.get('must_finish_by') else None,
                anticipated_start=datetime.strptime(request.form.get('anticipated_start'), '%Y-%m-%d') if request.form.get('anticipated_start') else None,
                anticipated_finish=datetime.strptime(request.form.get('anticipated_finish'), '%Y-%m-%d') if request.form.get('anticipated_finish') else None,
                expected_finish=datetime.strptime(request.form.get('expected_finish'), '%Y-%m-%d') if request.form.get('expected_finish') else None,
                data_date=datetime.utcnow()
            )
            db.session.add(dates)
            
            # إنشاء الميزانية
            budget = ProjectBudget(
                project_id=project.id,
                original_budget=float(request.form.get('original_budget', 0)),
                current_budget=float(request.form.get('original_budget', 0))
            )
            db.session.add(budget)
            
            # إنشاء التكاليف
            cost = ProjectCost(project_id=project.id)
            db.session.add(cost)
            
            # إنشاء الأداء
            performance = ProjectPerformance(project_id=project.id)
            db.session.add(performance)
            
            # إنشاء التقدم
            progress = ProjectProgress(project_id=project.id)
            db.session.add(progress)
            
            # إنشاء الإحصائيات
            statistics = ProjectStatistics(project_id=project.id)
            db.session.add(statistics)
            
            db.session.commit()
            
            # إنشاء إشعار للمدير
            if project.project_manager_id:
                notification = Notification(
                    user_id=project.project_manager_id,
                    title='تم تعيينك كمدير لمشروع جديد',
                    message=f'تم تعيينك كمدير لمشروع {project.name}',
                    notification_type='project_assigned',
                    related_project_id=project.id,
                )
                db.session.add(notification)
                db.session.commit()
            NotificationService.project_created(project, current_user)
            flash('تم إنشاء المشروع بنجاح', 'success')
            return redirect(url_for('projects.project_detail', project_id=project.id))
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error creating project: {str(e)}")
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    # بيانات النموذج للعرض
    projects_count = Project.query.filter_by(created_by=current_user.id).count()
    eps_list = EPS.query.filter_by(org_id=get_org_id()).all()
    obs_list = OBS.query.filter_by(org_id=get_org_id()).all()
    calendars = Calendar.query.filter_by(org_id=get_org_id()).all()
    users = User.query.filter_by(org_id=get_org_id()).all()
    clients = User.query.filter_by(org_id=get_org_id()).all()
    consultants = User.query.filter_by(org_id=get_org_id()).all()
    
    return render_template('projects/create.html',
                         projects_count=projects_count,
                         eps_list=eps_list,
                         obs_list=obs_list,
                         calendars=calendars,
                         users=users,
                         clients=clients,
                         eps_id=eps_id,
                         consultants=consultants,
                         now=datetime.now)

@project_bp.route('/api/check-code', methods=['POST'])
@login_required
def api_check_project_code():
    """API للتحقق من عدم تكرار كود المشروع"""
    data = request.get_json()
    project_code = data.get('project_code')
    
    existing = Project.query.filter_by(project_code=project_code).first()
    
    return jsonify({
        'success': True,
        'exists': existing is not None
    })
@project_bp.route('/<int:project_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_project(project_id):
    """تعديل مشروع"""
    project = check_project_access(project_id)
    if not project:
        flash('غير مصرح بالوصول', 'danger')
        return redirect(url_for('projects.list_projects'))
    
    if request.method == 'POST':
        try:
            # تحديث المعلومات الأساسية
            project.name = request.form.get('name', project.name)
            project.description = request.form.get('description', project.description)
            project.status = request.form.get('status', project.status)
            project.priority_level = int(request.form.get('priority_level', project.priority_level))
            project.risk_level = request.form.get('risk_level', project.risk_level)
            project.complexity = request.form.get('complexity', project.complexity)
            project.category = request.form.get('category', project.category)
            project.project_type = request.form.get('project_type', project.project_type)
            project.project_scale = request.form.get('project_scale', project.project_scale)
            project.website = request.form.get('website', project.website)
            
            # تحديث العلاقات
            project.eps_id = request.form.get('eps_id') or None
            project.obs_id = request.form.get('obs_id') or None
            project.calendar_id = request.form.get('calendar_id') or None
            project.financial_calendar_id = request.form.get('financial_calendar_id') or None
            project.project_manager_id = request.form.get('project_manager_id') or None
            project.checked_out_by = request.form.get('assistant_manager_id') or None
            project.client_id = request.form.get('client_id') or None
            project.consultant_id = request.form.get('consultant_id') or None
            project.base_currency = request.form.get('base_currency') or None
            # تحديث الموقع
            if project.location:
                project.location.site_name = request.form.get('site_name', project.location.site_name)
                project.location.city = request.form.get('city', project.location.city)
                project.location.country = request.form.get('country', project.location.country)
                project.location.address = request.form.get('address', project.location.address)
                project.location.coordinates = request.form.get('coordinates', project.location.coordinates)
            
            # تحديث التواريخ
            if project.dates:
                project.dates.planned_start = datetime.strptime(request.form.get('planned_start'), '%Y-%m-%d') if request.form.get('planned_start') else None
                project.dates.planned_finish = datetime.strptime(request.form.get('planned_finish'), '%Y-%m-%d') if request.form.get('planned_finish') else None
                project.dates.must_finish_by = datetime.strptime(request.form.get('must_finish_by'), '%Y-%m-%d') if request.form.get('must_finish_by') else None
                project.dates.anticipated_start = datetime.strptime(request.form.get('anticipated_start'), '%Y-%m-%d') if request.form.get('anticipated_start') else None
                project.dates.anticipated_finish = datetime.strptime(request.form.get('anticipated_finish'), '%Y-%m-%d') if request.form.get('anticipated_finish') else None
                project.dates.expected_finish = datetime.strptime(request.form.get('expected_finish'), '%Y-%m-%d') if request.form.get('expected_finish') else None
            
            # تحديث الميزانية
            if project.budget:
                project.budget.original_budget = float(request.form.get('original_budget', project.budget.original_budget))
                project.budget.current_budget = float(request.form.get('current_budget', project.budget.current_budget))
                project.budget.proposed_budget = float(request.form.get('proposed_budget', project.budget.proposed_budget))
            
            db.session.commit()
            flash('تم تحديث المشروع بنجاح', 'success')
            return redirect(url_for('projects.project_detail', project_id=project.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    # بيانات النموذج
    eps_list = EPS.query.filter_by(org_id=get_org_id()).all()
    obs_list = OBS.query.filter_by(org_id=get_org_id()).all()
    calendars = Calendar.query.filter_by(org_id=get_org_id()).all()
    users = User.query.filter_by(org_id=get_org_id()).all()
    clients = User.query.filter_by(org_id=get_org_id(),role="client").all()
    consultants = User.query.filter_by(org_id=get_org_id(),role="consultant").all()
    
    return render_template('projects/edit.html',
                         project=project,
                         eps_list=eps_list,
                         obs_list=obs_list,
                         calendars=calendars,
                         users=users,
                         consultants=consultants,
                         clients=clients)


@project_bp.route('/<int:project_id>/delete', methods=['POST'])
@login_required
def delete_project(project_id):
    """حذف مشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    try:
        db.session.delete(project)
        db.session.commit()
        flash('تم حذف المشروع بنجاح', 'success')
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================
# API Routes للمشاريع
# ============================================

@project_bp.route('/api/list')
@login_required
def api_project_list():
    """API لقائمة المشاريع"""
    projects = Project.query.filter_by(created_by=current_user.id).all()
    return jsonify({
        'success': True,
        'projects': [{
            'id': p.id,
            'name': p.name,
            'project_code': p.project_code,
            'status': p.status,
            'progress': p.progress.progress_percentage if p.progress else 0,
            'planned_finish': p.dates.planned_finish.isoformat() if p.dates and p.dates.planned_finish else None,
            'is_overdue': p.is_overdue
        } for p in projects]
    })


@project_bp.route('/api/<int:project_id>')
@login_required
def api_project_detail(project_id):
    """API لتفاصيل المشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    return jsonify({
        'success': True,
        'project': {
            'id': project.id,
            'name': project.name,
            'project_code': project.project_code,
            'description': project.description,
            'status': project.status,
            'priority_level': project.priority_level,
            'risk_level': project.risk_level,
            'complexity': project.complexity,
            'category': project.category,
            'project_type': project.project_type,
            'project_scale': project.project_scale,
            'location': {
                'site_name': project.location.site_name if project.location else None,
                'city': project.location.city if project.location else None,
                'country': project.location.country if project.location else None
            } if project.location else None,
            'dates': {
                'planned_start': project.dates.planned_start.isoformat() if project.dates and project.dates.planned_start else None,
                'planned_finish': project.dates.planned_finish.isoformat() if project.dates and project.dates.planned_finish else None,
                'actual_start': project.dates.actual_start.isoformat() if project.dates and project.dates.actual_start else None,
                'actual_finish': project.dates.actual_finish.isoformat() if project.dates and project.dates.actual_finish else None,
                'must_finish_by': project.dates.must_finish_by.isoformat() if project.dates and project.dates.must_finish_by else None,
                'expected_finish': project.dates.expected_finish.isoformat() if project.dates and project.dates.expected_finish else None
            } if project.dates else None,
            'budget': {
                'original': project.budget.original_budget if project.budget else 0,
                'current': project.budget.current_budget if project.budget else 0,
                'proposed': project.budget.proposed_budget if project.budget else 0
            } if project.budget else None,
            'progress': project.progress.progress_percentage if project.progress else 0,
            'statistics': {
                'total_activities': project.statistics.total_activities if project.statistics else 0,
                'completed_activities': project.statistics.completed_activities if project.statistics else 0,
                'total_tasks': project.statistics.total_tasks if project.statistics else 0,
                'completed_tasks': project.statistics.completed_tasks if project.statistics else 0,
                'critical_activities': project.statistics.critical_activities if project.statistics else 0
            } if project.statistics else None
        }
    })

@project_bp.route('/start/<int:project_id>', methods=['POST'])
@login_required
def start_project(project_id):
    """بدء المشروع"""
    project = Project.query.get_or_404(project_id)
    
    # التحقق من الصلاحية
    if project.project_manager_id != current_user.id and current_user.role != 'org_admin':
        return jsonify({'error': _('access_denied')}), 403
    
    # التحقق من الحالة الحالية
    if project.status == 'in_progress':
        return jsonify({'error': _('project_already_started')}), 400
    
    if project.status == 'completed':
        return jsonify({'error': _('project_already_completed')}), 400
    
    try:
        # تحديث حالة المشروع
        project.status = 'in_progress'
        if not project.dates:
            project.dates = ProjectDates(project_id=project.id)
        project.dates.actual_start = datetime.utcnow()
        project.dates.data_date = datetime.utcnow()
        
        db.session.commit()
        
        # ✅ تحديث المؤشرات
        UpdateService.update_project_metrics(project)
        
        # تسجيل التقدم الأولي
        try:
            monitor = SmartMonitoringSystem()
            monitor.record_project_progress(project)
        except Exception as e:
            logger.warning(f"خطأ في تسجيل التقدم: {str(e)}")
        
        # إرسال إشعارات
        try:
            NotificationService.project_started(project)
        except Exception as e:
            logger.warning(f"خطأ في إرسال الإشعارات: {str(e)}")
        
        return jsonify({
            'success': True, 
            'message': _('project_started_success'),
            'status': project.status
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"خطأ في بدء المشروع: {str(e)}")
        return jsonify({'error': str(e)}), 500

    
@project_bp.route('/complete/<int:project_id>', methods=['POST'])
@login_required
def complete_project(project_id):
    """إكمال المشروع"""
    project = Project.query.get_or_404(project_id)
    
    # التحقق من الصلاحية
    if project.project_manager_id != current_user.id and current_user.role != 'org_admin':
        return jsonify({'error': _('access_denied')}), 403
    
    # التحقق من أن المشروع قد بدأ بالفعل
    if project.status != 'in_progress':
        return jsonify({'error': _('project_not_started')}), 400
    
    try:
        # ============================================
        # التحقق من اكتمال جميع الأنشطة والمهام
        # ============================================
        
        # 1. التحقق من الأنشطة
        activities = Activity.query.filter_by(project_id=project_id).all()
        incomplete_activities = [a for a in activities if a.status != 'completed']
        
        # 2. التحقق من المهام
        tasks = Task.query.filter_by(project_id=project_id).all()
        incomplete_tasks = [t for t in tasks if t.status != 'completed']
        
        # 3. التحقق من المعالم
        milestones = Milestone.query.filter_by(project_id=project_id).all()
        incomplete_milestones = [m for m in milestones if m.status != 'achieved']
        
        # 4. التحقق من القضايا المفتوحة
        open_issues = Issue.query.filter_by(project_id=project_id).filter(Issue.status.in_(['open', 'in_progress'])).all()
        
        # تجميع الأخطاء
        errors = []
        
        if incomplete_activities:
            errors.append({
                'type': 'activities',
                'count': len(incomplete_activities),
                'message': _('incomplete_activities_count', count=len(incomplete_activities)),
                'items': [{'id': a.id, 'name': a.activity_name} for a in incomplete_activities[:5]]
            })
        
        if incomplete_tasks:
            errors.append({
                'type': 'tasks',
                'count': len(incomplete_tasks),
                'message': _('incomplete_tasks_count', count=len(incomplete_tasks)),
                'items': [{'id': t.id, 'name': t.task_name} for t in incomplete_tasks[:5]]
            })
        
        if incomplete_milestones:
            errors.append({
                'type': 'milestones',
                'count': len(incomplete_milestones),
                'message': _('incomplete_milestones_count', count=len(incomplete_milestones)),
                'items': [{'id': m.id, 'name': m.name} for m in incomplete_milestones[:5]]
            })
        
        if open_issues:
            errors.append({
                'type': 'issues',
                'count': len(open_issues),
                'message': _('open_issues_count', count=len(open_issues)),
                'items': [{'id': i.id, 'title': i.title} for i in open_issues[:5]]
            })
        
        # إذا كان هناك عناصر غير مكتملة، أعد خطأ مع التفاصيل
        if errors:
            return jsonify({
                'success': False,
                'error': _('cannot_complete_project'),
                'details': errors,
                'can_force': current_user.role == 'org_admin'
            }), 400
        
        # ============================================
        # إكمال المشروع
        # ============================================
        
        # تحديث حالة المشروع
        project.status = 'completed'
        if not project.dates:
            project.dates = ProjectDates(project_id=project.id)
        project.dates.actual_finish = datetime.utcnow()
        
        # تحديث حالة الأنشطة غير المكتملة (إذا وجدت)
        for activity in incomplete_activities:
            if activity.status != 'completed':
                activity.status = 'completed'
                activity.actual_finish = datetime.utcnow()
                activity.progress_percentage = 100
        
        # تحديث حالة المهام غير المكتملة
        for task in incomplete_tasks:
            if task.status != 'completed':
                task.status = 'completed'
                if not task.execution:
                    task.execution = TaskExecution(task_id=task.id)
                task.execution.actual_finish = datetime.utcnow()
                if task.progress:
                    task.progress.progress_percentage = 100
        
        # تحديث حالة المعالم غير المحققة
        for milestone in incomplete_milestones:
            if milestone.status != 'achieved':
                milestone.status = 'achieved'
                milestone.actual_date = datetime.now().date()
        
        db.session.commit()
        
        # ✅ تحديث المؤشرات النهائية
        UpdateService.update_project_metrics(project)
        
        # تسجيل التقدم النهائي
        try:
            monitor = SmartMonitoringSystem()
            monitor.record_project_progress(project)
        except Exception as e:
            logger.warning(f"خطأ في تسجيل التقدم: {str(e)}")
        
        # إرسال إشعارات
        try:
            NotificationService.project_completed(project)
        except Exception as e:
            logger.warning(f"خطأ في إرسال الإشعارات: {str(e)}")
        
        return jsonify({
            'success': True, 
            'message': _('project_completed_success'),
            'status': project.status
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"خطأ في إكمال المشروع: {str(e)}")
        return jsonify({'error': str(e)}), 500


@project_bp.route('/<int:project_id>/check-completion', methods=['GET'])
@login_required
def check_project_completion(project_id):
    """التحقق من جاهزية المشروع للإكمال"""
    project = Project.query.get_or_404(project_id)
    
    # التحقق من الصلاحية
    if project.project_manager_id != current_user.id and current_user.role != 'org_admin':
        return jsonify({'error': _('access_denied')}), 403
    
    # جمع إحصائيات الإكمال
    activities = Activity.query.filter_by(project_id=project_id).all()
    tasks = Task.query.filter_by(project_id=project_id).all()
    milestones = Milestone.query.filter_by(project_id=project_id).all()
    open_issues = Issue.query.filter_by(project_id=project_id).filter(Issue.status.in_(['open', 'in_progress'])).all()
    
    stats = {
        'total_activities': len(activities),
        'completed_activities': len([a for a in activities if a.status == 'completed']),
        'total_tasks': len(tasks),
        'completed_tasks': len([t for t in tasks if t.status == 'completed']),
        'total_milestones': len(milestones),
        'completed_milestones': len([m for m in milestones if m.status == 'achieved']),
        'open_issues': len(open_issues),
        'can_complete': (
            len([a for a in activities if a.status != 'completed']) == 0 and
            len([t for t in tasks if t.status != 'completed']) == 0 and
            len([m for m in milestones if m.status != 'achieved']) == 0 and
            len(open_issues) == 0
        )
    }
    
    stats['completion_percentage'] = (
        (stats['completed_activities'] + stats['completed_tasks'] + stats['completed_milestones']) /
        (stats['total_activities'] + stats['total_tasks'] + stats['total_milestones']) * 100
    ) if (stats['total_activities'] + stats['total_tasks'] + stats['total_milestones']) > 0 else 0
    
    return jsonify({'success': True, 'stats': stats})

@project_bp.route('/<int:project_id>/check-completion-details', methods=['GET'])
@login_required
def check_completion_details(project_id):
    """جلب تفاصيل العناصر غير المكتملة للمشروع"""
    project = Project.query.get_or_404(project_id)
    
    if project.project_manager_id != current_user.id and current_user.role != 'org_admin':
        return jsonify({'error': _('access_denied')}), 403
    
    incomplete_activities = Activity.query.filter(
        Activity.project_id == project_id,
        Activity.status != 'completed'
    ).all()
    
    incomplete_tasks = Task.query.filter(
        Task.project_id == project_id,
        Task.status != 'completed'
    ).all()
    
    incomplete_milestones = Milestone.query.filter(
        Milestone.project_id == project_id,
        Milestone.status != 'achieved'
    ).all()
    
    open_issues = Issue.query.filter(
        Issue.project_id == project_id,
        Issue.status.in_(['open', 'in_progress'])
    ).all()
    
    return jsonify({
        'success': True,
        'incomplete_activities': [{'id': a.id, 'name': a.activity_name, 'status': a.status} for a in incomplete_activities],
        'incomplete_tasks': [{'id': t.id, 'name': t.task_name, 'status': t.status} for t in incomplete_tasks],
        'incomplete_milestones': [{'id': m.id, 'name': m.name, 'status': m.status} for m in incomplete_milestones],
        'open_issues': [{'id': i.id, 'title': i.title, 'status': i.status} for i in open_issues]
    })

@project_bp.route('/force-complete/<int:project_id>', methods=['POST'])
@login_required
def force_complete_project(project_id):
    """إكمال المشروع قسرياً (للمسؤول فقط)"""
    project = Project.query.get_or_404(project_id)
    
    # فقط مدير المؤسسة يمكنه الإكمال القسري
    if current_user.role != 'org_admin':
        return jsonify({'error': _('access_denied')}), 403
    
    try:
        project.status = 'completed'
        if not project.dates:
            project.dates = ProjectDates(project_id=project.id)
        project.dates.actual_finish = datetime.utcnow()
        
        db.session.commit()
        
        UpdateService.update_project_metrics(project)
        
        return jsonify({
            'success': True,
            'message': _('project_force_completed_success'),
            'status': project.status
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    
@project_bp.route('/api/<int:project_id>/update', methods=['POST'])
@login_required
def api_update_project(project_id):
    """تحديث بيانات المشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        if 'name' in data:
            project.name = data['name']
        if 'description' in data:
            project.description = data['description']
        if 'status' in data:
            project.status = data['status']
        if 'priority_level' in data:
            project.priority_level = data['priority_level']
        if 'risk_level' in data:
            project.risk_level = data['risk_level']
        
        if 'dates' in data and project.dates:
            if 'planned_start' in data['dates']:
                project.dates.planned_start = datetime.strptime(data['dates']['planned_start'], '%Y-%m-%d') if data['dates']['planned_start'] else None
            if 'planned_finish' in data['dates']:
                project.dates.planned_finish = datetime.strptime(data['dates']['planned_finish'], '%Y-%m-%d') if data['dates']['planned_finish'] else None
        
        if 'budget' in data and project.budget:
            if 'current_budget' in data['budget']:
                project.budget.current_budget = float(data['budget']['current_budget'])
                # ✅ تحديث المؤشرات (القيمة المكتسبة تتغير)
                UpdateService.update_project_metrics(project)
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================
# API Routes للميزانية والتمويل
# ============================================

@project_bp.route('/api/<int:project_id>/budget-logs', methods=['GET'])
@login_required
def api_budget_logs(project_id):
    """API لجلب سجل الميزانية"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    logs = BudgetLog.query.filter_by(project_id=project_id).order_by(BudgetLog.date.desc()).all()
    
    return jsonify({
        'success': True,
        'logs': [{
            'id': l.id,
            'date': l.date.isoformat(),
            'change_number': l.change_number,
            'amount': l.amount,
            'responsible': l.responsible.full_name if l.responsible else None,
            'status': l.status,
            'reason': l.reason
        } for l in logs]
    })


@project_bp.route('/api/<int:project_id>/budget-log', methods=['POST'])
@login_required
def api_add_budget_log(project_id):
    """API لإضافة سجل ميزانية"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        last_log = BudgetLog.query.filter_by(project_id=project_id).order_by(BudgetLog.id.desc()).first()
        if last_log and last_log.change_number:
            last_num = int(last_log.change_number.split('-')[-1])
            change_number = f"CHG-{datetime.now().strftime('%Y%m')}-{last_num + 1:04d}"
        else:
            change_number = f"CHG-{datetime.now().strftime('%Y%m')}-0001"
        
        log = BudgetLog(
            project_id=project_id,
            date=datetime.strptime(data['date'], '%Y-%m-%d').date() if data.get('date') else date.today(),
            change_number=change_number,
            amount=float(data.get('amount', 0)),
            responsible_id=current_user.id,
            status=data.get('status', 'Proposed'),
            reason=data.get('reason')
        )
        db.session.add(log)
        
        if data.get('status') == 'Approved' and project.budget:
            project.budget.current_budget = (project.budget.current_budget or 0) + float(data.get('amount', 0))
        
        db.session.commit()
        return jsonify({'success': True, 'log_id': log.id})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/<int:project_id>/funding', methods=['GET', 'POST'])
@login_required
def api_funding(project_id):
    """API لإدارة مصادر التمويل"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    if request.method == 'GET':
        sources = FundingSource.query.filter_by(project_id=project_id).all()
        return jsonify({
            'success': True,
            'sources': [{
                'id': s.id,
                'source_name': s.source_name,
                'amount': s.amount,
                'share_percentage': s.share_percentage,
                'currency': s.currency,
                'status': s.status
            } for s in sources],
            'total': sum(s.amount for s in sources)
        })
    
    else:  # POST
        data = request.get_json()
        try:
            source = FundingSource(
                project_id=project_id,
                source_name=data['source_name'],
                amount=float(data.get('amount', 0)),
                share_percentage=float(data.get('share_percentage', 0)),
                currency=data.get('currency', 'SAR'),
                status=data.get('status', 'Proposed')
            )
            db.session.add(source)
            db.session.commit()
            return jsonify({'success': True, 'source_id': source.id})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/<int:project_id>/spending-plan', methods=['GET', 'POST'])
@login_required
def api_spending_plan(project_id):
    """API لإدارة خطة الصرف"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    if request.method == 'GET':
        items = SpendingPlanItem.query.filter_by(project_id=project_id).order_by(SpendingPlanItem.date).all()
        
        # حساب المجاميع التراكمية
        running_spending = 0
        running_benefit = 0
        for item in items:
            running_spending += item.planned_amount
            running_benefit += item.benefit_amount
            item.spending_tally = running_spending
            item.benefit_tally = running_benefit
            item.variance = item.benefit_amount - item.planned_amount
        
        return jsonify({
            'success': True,
            'items': [{
                'id': i.id,
                'date': i.date.isoformat(),
                'planned_amount': i.planned_amount,
                'benefit_amount': i.benefit_amount,
                'actual_amount': i.actual_amount,
                'spending_tally': i.spending_tally,
                'benefit_tally': i.benefit_tally,
                'variance': i.variance
            } for i in items],
            'totals': {
                'spending': sum(i.planned_amount for i in items),
                'benefit': sum(i.benefit_amount for i in items),
                'actual': sum(i.actual_amount for i in items)
            }
        })
    
    else:  # POST
        data = request.get_json()
        try:
            item = SpendingPlanItem(
                project_id=project_id,
                date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
                planned_amount=float(data.get('planned_amount', 0)),
                benefit_amount=float(data.get('benefit_amount', 0))
            )
            db.session.add(item)
            db.session.commit()
            return jsonify({'success': True, 'item_id': item.id})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

# ============================================
# API Routes للمعالم (Milestones)
# ============================================

@project_bp.route('/api/<int:project_id>/milestones', methods=['GET'])
@login_required
def api_milestones(project_id):
    """API لجلب معالم المشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    milestones = Milestone.query.filter_by(project_id=project_id).order_by(Milestone.planned_date).all()
    
    return jsonify({
        'success': True,
        'milestones': [{
            'id': m.id,
            'code': m.milestone_code,
            'name': m.name,
            'planned_date': m.planned_date.isoformat(),
            'actual_date': m.actual_date.isoformat() if m.actual_date else None,
            'type': m.milestone_type,
            'status': m.status,
            'weight': m.weight
        } for m in milestones]
    })


@project_bp.route('/api/<int:project_id>/milestone', methods=['POST'])
@login_required
def api_add_milestone(project_id):
    """API لإضافة معلم جديد"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        milestone = Milestone(
            project_id=project_id,
            milestone_code=data.get('code'),
            name=data.get('name'),
            description=data.get('description'),
            planned_date=datetime.strptime(data['planned_date'], '%Y-%m-%d').date(),
            milestone_type=data.get('type'),
            weight=float(data.get('weight', 0)),
            status='pending'
        )
        
        db.session.add(milestone)
        db.session.commit()
        
        return jsonify({'success': True, 'milestone_id': milestone.id})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/milestone/<int:milestone_id>/update-status', methods=['POST'])
@login_required
def api_update_milestone_status(milestone_id):
    """API لتحديث حالة معلم"""
    milestone = Milestone.query.get_or_404(milestone_id)
    
    project = Project.query.get(milestone.project_id)
    if project.created_by != current_user.id and current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        milestone.status = data.get('status', milestone.status)
        if data.get('status') == 'achieved':
            milestone.actual_date = date.today()
            milestone.achieved_by = current_user.id
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================
# API Routes للأكواد والحقول المخصصة
# ============================================


@project_bp.route('/api/<int:project_id>/udf', methods=['GET', 'POST'])
@login_required
def api_project_udf(project_id):
    """API لإدارة الحقول المخصصة"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    if request.method == 'GET':
        udfs = ProjectUDF.query.filter_by(project_id=project_id).all()
        return jsonify({
            'success': True,
            'udfs': [{
                'id': u.id,
                'udf_name': u.udf_name,
                'udf_value': u.udf_value,
                'udf_type': u.udf_type
            } for u in udfs]
        })
    
    else:  # POST
        data = request.get_json()
        try:
            udf = ProjectUDF.query.filter_by(
                project_id=project_id,
                udf_name=data['udf_name']
            ).first()
            
            if udf:
                udf.udf_value = data['udf_value']
                udf.udf_type = data.get('udf_type', 'text')
            else:
                udf = ProjectUDF(
                    project_id=project_id,
                    udf_name=data['udf_name'],
                    udf_value=data['udf_value'],
                    udf_type=data.get('udf_type', 'text')
                )
                db.session.add(udf)
            
            db.session.commit()
            return jsonify({'success': True})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

# ============================================
# API Routes للإحصائيات
# ============================================

@project_bp.route('/api/<int:project_id>/statistics')
@login_required
def api_project_statistics(project_id):
    """API للحصول على إحصائيات المشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    if project.statistics:
        project.statistics.update()
        db.session.commit()
    
    stats = project.statistics
    
    return jsonify({
        'success': True,
        'statistics': {
            'total_activities': stats.total_activities if stats else 0,
            'completed_activities': stats.completed_activities if stats else 0,
            'in_progress_activities': stats.in_progress_activities if stats else 0,
            'not_started_activities': stats.not_started_activities if stats else 0,
            'critical_activities': stats.critical_activities if stats else 0,
            'total_tasks': stats.total_tasks if stats else 0,
            'completed_tasks': stats.completed_tasks if stats else 0,
            'total_resources': stats.total_resources if stats else 0,
            'total_manpower': stats.total_manpower if stats else 0
        }
    })


@project_bp.route('/api/<int:project_id>/performance')
@login_required
def api_project_performance(project_id):
    """API للحصول على أداء المشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    perf = project.performance
    
    return jsonify({
        'success': True,
        'performance': {
            'spi': perf.spi if perf else 1.0,
            'cpi': perf.cpi if perf else 1.0,
            'csi': perf.csi if perf else 1.0,
            'eac': perf.eac if perf else 0,
            'etc': perf.etc if perf else 0,
            'vac': perf.vac if perf else 0,
            'planned_value': perf.planned_value if perf else 0,
            'earned_value': perf.earned_value if perf else 0,
            'actual_cost': perf.actual_cost if perf else 0
        } if perf else {}
    })

# ============================================
# API Routes للعملاء والموردين
# ============================================

@project_bp.route('/api/clients')
@login_required
def api_clients_list():
    """API لقائمة العملاء"""
    clients = Client.query.filter_by(org_id=get_org_id()).all()
    
    return jsonify({
        'success': True,
        'clients': [{
            'id': c.id,
            'name': c.name,
            'code': c.client_code,
            'type': c.type,
            'contact': c.contact_person,
            'phone': c.phone,
            'email': c.email
        } for c in clients]
    })


@project_bp.route('/api/consultants')
@login_required
def api_consultants_list():
    """API لقائمة الاستشاريين"""
    consultants = Consultant.query.filter_by(org_id=get_org_id()).all()
    
    return jsonify({
        'success': True,
        'consultants': [{
            'id': c.id,
            'name': c.name,
            'code': c.consultant_code,
            'specialization': c.specialization,
            'contact': c.contact_person,
            'phone': c.phone,
            'email': c.email
        } for c in consultants]
    })


@project_bp.route('/api/suppliers')
@login_required
def api_suppliers_list():
    """API لقائمة الموردين"""
    suppliers = Supplier.query.filter_by(org_id=get_org_id()).all()
    
    return jsonify({
        'success': True,
        'suppliers': [{
            'id': s.id,
            'name': s.name,
            'code': s.supplier_code,
            'type': s.type,
            'contact': s.contact_person,
            'phone': s.phone,
            'email': s.email,
            'rating': s.rating,
            'is_approved': s.is_approved
        } for s in suppliers]
    })
@project_bp.route('/api/<int:project_id>/update-dates', methods=['POST'])
@login_required
def api_update_project_dates(project_id):
    """تحديث تواريخ المشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        if project.dates:
            if 'planned_start' in data and data['planned_start']:
                project.dates.planned_start = datetime.strptime(data['planned_start'], '%Y-%m-%d')
            if 'must_finish_by' in data and data['must_finish_by']:
                project.dates.must_finish_by = datetime.strptime(data['must_finish_by'], '%Y-%m-%d')
            if 'data_date' in data and data['data_date']:
                project.dates.data_date = datetime.strptime(data['data_date'], '%Y-%m-%d')
            if 'anticipated_start' in data and data['anticipated_start']:
                project.dates.anticipated_start = datetime.strptime(data['anticipated_start'], '%Y-%m-%d')
            if 'anticipated_finish' in data and data['anticipated_finish']:
                project.dates.anticipated_finish = datetime.strptime(data['anticipated_finish'], '%Y-%m-%d')
            if 'actual_start' in data and data['actual_start']:
                project.dates.actual_start = datetime.strptime(data['actual_start'], '%Y-%m-%d')
            if 'actual_finish' in data and data['actual_finish']:
                project.dates.actual_finish = datetime.strptime(data['actual_finish'], '%Y-%m-%d')
            if 'baseline_start' in data and data['baseline_start']:
                project.dates.baseline_start = datetime.strptime(data['baseline_start'], '%Y-%m-%d')
            if 'baseline_finish' in data and data['baseline_finish']:
                project.dates.baseline_finish = datetime.strptime(data['baseline_finish'], '%Y-%m-%d')
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================
# API Routes للملاحظات (Notebook)
# ============================================

@project_bp.route('/api/<int:project_id>/notebook/entries', methods=['GET'])
@login_required
def api_notebook_entries(project_id):
    """جلب جميع ملاحظات المشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    entries = NotebookEntry.query.filter_by(project_id=project_id).order_by(NotebookEntry.created_at.desc()).all()
    
    return jsonify({
        'success': True,
        'entries': [{
            'id': e.id,
            'topic': e.topic,
            'content': e.content,
            'author': e.creator.full_name if e.creator else None,
            'date': e.created_at.strftime('%Y-%m-%d %H:%M'),
            'preview': e.content[:100] + '...' if len(e.content) > 100 else e.content
        } for e in entries]
    })


@project_bp.route('/api/<int:project_id>/notebook/entry', methods=['POST'])
@login_required
def api_add_notebook_entry(project_id):
    """إضافة ملاحظة جديدة"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        entry = NotebookEntry(
            project_id=project_id,
            topic=data.get('topic', 'Comments'),
            content=data['content'],
            created_by=current_user.id
        )
        
        db.session.add(entry)
        db.session.commit()
        
        return jsonify({'success': True, 'entry_id': entry.id})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# API Routes لخطة الصرف (Spending Plan)
# ============================================

# @project_bp.route('/api/<int:project_id>/spending-plan', methods=['GET'])
# @login_required
# def api_spending_plan(project_id):
#     """جلب خطة الصرف"""
#     project = check_project_access(project_id)
#     if not project:
#         return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
#     items = SpendingPlanItem.query.filter_by(project_id=project_id).order_by(SpendingPlanItem.date).all()
    
#     # حساب المجاميع التراكمية
#     running_spending = 0
#     running_benefit = 0
    
#     for item in items:
#         running_spending += item.planned_amount
#         running_benefit += item.benefit_amount
#         item.spending_tally = running_spending
#         item.benefit_tally = running_benefit
#         item.undistributed_variance = running_spending - running_benefit
#         item.benefit_variance = item.benefit_amount - item.planned_amount
    
#     return jsonify({
#         'success': True,
#         'items': [{
#             'id': i.id,
#             'date': i.date.strftime('%Y-%m-%d'),
#             'planned_amount': i.planned_amount,
#             'benefit_amount': i.benefit_amount,
#             'actual_amount': i.actual_amount,
#             'spending_tally': i.spending_tally,
#             'benefit_tally': i.benefit_tally,
#             'undistributed_variance': i.undistributed_variance,
#             'benefit_variance': i.benefit_variance
#         } for i in items],
#         'totals': {
#             'spending': sum(i.planned_amount for i in items),
#             'benefit': sum(i.benefit_amount for i in items),
#             'actual': sum(i.actual_amount for i in items)
#         }
#     })


@project_bp.route('/api/<int:project_id>/spending-item', methods=['POST'])
@login_required
def api_add_spending_item(project_id):
    """إضافة بند خطة صرف"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        item = SpendingPlanItem(
            project_id=project_id,
            date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
            planned_amount=float(data.get('planned_amount', 0)),
            benefit_amount=float(data.get('benefit_amount', 0))
        )
        
        db.session.add(item)
        db.session.commit()
        
        return jsonify({'success': True, 'item_id': item.id})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@project_bp.route('/api/spending-item/<int:item_id>/update', methods=['POST'])
@login_required
def api_update_spending_item(item_id):
    """تحديث بند خطة الصرف"""
    item = SpendingPlanItem.query.get_or_404(item_id)
    
    project = Project.query.get(item.project_id)
    if project.created_by != current_user.id and current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        if 'date' in data:
            item.date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        if 'planned_amount' in data:
            item.planned_amount = float(data['planned_amount'])
        if 'benefit_amount' in data:
            item.benefit_amount = float(data['benefit_amount'])
        
        # إعادة حساب المجاميع التراكمية
        items = SpendingPlanItem.query.filter_by(project_id=item.project_id).order_by(SpendingPlanItem.date).all()
        
        running_spending = 0
        running_benefit = 0
        
        for i in items:
            running_spending += i.planned_amount
            running_benefit += i.benefit_amount
            i.spending_tally = running_spending
            i.benefit_tally = running_benefit
            i.undistributed_variance = running_spending - running_benefit
            i.benefit_variance = i.benefit_amount - i.planned_amount
        
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/spending-item/<int:item_id>/delete', methods=['POST'])
@login_required
def api_delete_spending_item(item_id):
    """حذف بند خطة الصرف"""
    item = SpendingPlanItem.query.get_or_404(item_id)
    
    project = Project.query.get(item.project_id)
    if project.created_by != current_user.id and current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    try:
        db.session.delete(item)
        
        # إعادة حساب المجاميع التراكمية للعناصر المتبقية
        remaining_items = SpendingPlanItem.query.filter_by(project_id=item.project_id).order_by(SpendingPlanItem.date).all()
        
        running_spending = 0
        running_benefit = 0
        
        for i in remaining_items:
            running_spending += i.planned_amount
            running_benefit += i.benefit_amount
            i.spending_tally = running_spending
            i.benefit_tally = running_benefit
            i.undistributed_variance = running_spending - running_benefit
            i.benefit_variance = i.benefit_amount - i.planned_amount
        
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/<int:project_id>/spending-plan/summary')
@login_required
def api_spending_plan_summary(project_id):
    """الحصول على ملخص خطة الصرف"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    items = SpendingPlanItem.query.filter_by(project_id=project_id).order_by(SpendingPlanItem.date).all()
    
    if not items:
        return jsonify({
            'success': True,
            'summary': {
                'total_spending': 0,
                'total_benefit': 0,
                'undistributed_variance': 0,
                'benefit_variance': 0,
                'last_tally': 0
            }
        })
    
    total_spending = sum(i.planned_amount for i in items)
    total_benefit = sum(i.benefit_amount for i in items)
    
    # الحصول على آخر العناصر للحصول على المجاميع التراكمية
    last_item = items[-1]
    
    return jsonify({
        'success': True,
        'summary': {
            'total_spending': total_spending,
            'total_benefit': total_benefit,
            'undistributed_variance': last_item.undistributed_variance if hasattr(last_item, 'undistributed_variance') else total_spending - total_benefit,
            'benefit_variance': total_benefit - total_spending,
            'spending_tally': last_item.spending_tally if hasattr(last_item, 'spending_tally') else total_spending,
            'benefit_tally': last_item.benefit_tally if hasattr(last_item, 'benefit_tally') else total_benefit
        }
    })
# ============================================
# API Routes للأكواد (Codes)
# ============================================

# @project_bp.route('/api/<int:project_id>/codes', methods=['GET'])
# @login_required
# def api_project_codes(project_id):
#     """جلب أكواد المشروع"""
#     project = check_project_access(project_id)
#     if not project:
#         return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
#     codes = ProjectCode.query.filter_by(project_id=project_id).all()
    
#     return jsonify({
#         'success': True,
#         'codes': [{
#             'id': c.id,
#             'code_type': c.code_type,
#             'code_value': c.code_value,
#             'description': c.description
#         } for c in codes]
#     })
@project_bp.route('/api/<int:project_id>/codes', methods=['GET'])
@login_required
def api_get_project_codes(project_id):
    """API لجلب جميع أكواد المشروع (حسب النظام الجديد)"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    try:
        # الحصول على جميع قواميس الأكواد النشطة للمؤسسة
        dictionaries = ProjectCodeDictionary.query.filter_by(
            org_id=project.org_id,
            is_active=True
        ).all()
        
        result = []
        for dictionary in dictionaries:
            # البحث عن التعيين لهذا المشروع
            assignment = ProjectCodeAssignment.query.filter_by(
                project_id=project_id,
                dictionary_id=dictionary.id
            ).first()
            
            # بناء هيكل الأكواد للقاموس
            def build_tree(parent_id=None):
                tree = []
                codes = ProjectCodeValue.query.filter_by(
                    dictionary_id=dictionary.id,
                    parent_id=parent_id,
                    is_active=True
                ).order_by(ProjectCodeValue.display_sequence).all()
                
                for code in codes:
                    children = build_tree(code.id)
                    code_dict = {
                        'id': code.id,
                        'code_value': code.code_value,
                        'code_value_ar': code.code_value_ar,
                        'description': code.code_description,
                        'display_color': code.display_color,
                        'level': code.level,
                        'full_path': code.full_path,
                        'has_children': len(children) > 0
                    }
                    if children:
                        code_dict['children'] = children
                    tree.append(code_dict)
                return tree
            
            result.append({
                'dictionary_id': dictionary.id,
                'dictionary_name': dictionary.dict_name,
                'dictionary_name_ar': dictionary.dict_name_ar,
                'dictionary_description': dictionary.description,
                'is_hierarchical': dictionary.is_hierarchical,
                'assigned_code_id': assignment.code_value_id if assignment else None,
                'assigned_code': {
                    'id': assignment.code_value.id,
                    'code_value': assignment.code_value.code_value,
                    'code_value_ar': assignment.code_value.code_value_ar,
                    'description': assignment.code_value.code_description,
                    'display_color': assignment.code_value.display_color
                } if assignment else None,
                'available_codes': build_tree()
            })
        
        return jsonify({
            'success': True,
            'dictionaries': result
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/<int:project_id>/codes/assign', methods=['POST'])
@login_required
def api_assign_project_code(project_id):
    """API لتعيين كود لمشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    dictionary_id = data.get('dictionary_id')
    code_value_id = data.get('code_value_id')
    
    if not dictionary_id or not code_value_id:
        return jsonify({'success': False, 'error': 'البيانات غير كاملة'}), 400
    
    try:
        # البحث عن تعيين موجود
        assignment = ProjectCodeAssignment.query.filter_by(
            project_id=project_id,
            dictionary_id=dictionary_id
        ).first()
        
        if assignment:
            # تحديث التعيين الموجود
            assignment.code_value_id = code_value_id
            assignment.created_by = current_user.id
        else:
            # إنشاء تعيين جديد
            assignment = ProjectCodeAssignment(
                project_id=project_id,
                dictionary_id=dictionary_id,
                code_value_id=code_value_id,
                created_by=current_user.id
            )
            db.session.add(assignment)
        
        db.session.commit()
        
        # جلب معلومات الكود المعين للرد
        code_value = ProjectCodeValue.query.get(code_value_id)
        
        return jsonify({
            'success': True,
            'assigned_code': {
                'id': code_value.id,
                'code_value': code_value.code_value,
                'code_value_ar': code_value.code_value_ar,
                'display_color': code_value.display_color
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/<int:project_id>/codes/unassign', methods=['POST'])
@login_required
def api_unassign_project_code(project_id):
    """API لإزالة كود من مشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    dictionary_id = data.get('dictionary_id')
    
    if not dictionary_id:
        return jsonify({'success': False, 'error': 'معرف القاموس مطلوب'}), 400
    
    try:
        assignment = ProjectCodeAssignment.query.filter_by(
            project_id=project_id,
            dictionary_id=dictionary_id
        ).first()
        
        if assignment:
            db.session.delete(assignment)
            db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/<int:project_id>/codes/remove', methods=['POST'])
@login_required
def api_remove_project_codes(project_id):
    """API لإزالة عدة أكواد من مشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    dictionary_ids = data.get('dictionary_ids', [])
    
    if not dictionary_ids:
        return jsonify({'success': False, 'error': 'لا توجد أكواد محددة'}), 400
    
    try:
        # حذف التعيينات للقواميس المحددة
        ProjectCodeAssignment.query.filter(
            ProjectCodeAssignment.project_id == project_id,
            ProjectCodeAssignment.dictionary_id.in_(dictionary_ids)
        ).delete(synchronize_session=False)
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/dictionaries', methods=['GET'])
@login_required
def api_get_code_dictionaries():
    """API لجلب جميع قواميس الأكواد للمؤسسة"""
    org_id = current_user.org_id
    
    dictionaries = ProjectCodeDictionary.query.filter_by(
        org_id=org_id,
        is_active=True
    ).order_by(ProjectCodeDictionary.dict_name).all()
    
    return jsonify({
        'success': True,
        'dictionaries': [{
            'id': d.id,
            'name': d.dict_name,
            'name_ar': d.dict_name_ar,
            'description': d.description,
            'is_hierarchical': d.is_hierarchical,
            'codes_count': d.codes.count()
        } for d in dictionaries]
    })


@project_bp.route('/api/dictionary/<int:dictionary_id>/codes', methods=['GET'])
@login_required
def api_get_dictionary_codes(dictionary_id):
    """API لجلب جميع قيم الأكواد لقاموس معين"""
    dictionary = ProjectCodeDictionary.query.get_or_404(dictionary_id)
    
    if dictionary.org_id != current_user.org_id:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    # بناء الشجرة
    def build_tree(parent_id=None):
        tree = []
        codes = ProjectCodeValue.query.filter_by(
            dictionary_id=dictionary_id,
            parent_id=parent_id,
            is_active=True
        ).order_by(ProjectCodeValue.display_sequence).all()
        
        for code in codes:
            children = build_tree(code.id)
            code_dict = {
                'id': code.id,
                'code_value': code.code_value,
                'code_value_ar': code.code_value_ar,
                'description': code.code_description,
                'display_color': code.display_color,
                'display_sequence': code.display_sequence,
                'level': code.level,
                'full_path': code.full_path,
                'has_children': len(children) > 0,
                'assignments_count': code.assignments.count()
            }
            if children:
                code_dict['children'] = children
            tree.append(code_dict)
        return tree
    
    return jsonify({
        'success': True,
        'dictionary': {
            'id': dictionary.id,
            'name': dictionary.dict_name,
            'name_ar': dictionary.dict_name_ar,
            'description': dictionary.description,
            'is_hierarchical': dictionary.is_hierarchical
        },
        'codes': build_tree()
    })


@project_bp.route('/api/code/<int:code_id>', methods=['GET'])
@login_required
def api_get_code_details(code_id):
    """API لجلب تفاصيل قيمة كود محددة"""
    code = ProjectCodeValue.query.get_or_404(code_id)
    
    if code.dictionary.org_id != current_user.org_id:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    # الحصول على المسار الكامل
    path_parts = []
    current = code
    while current:
        path_parts.insert(0, current.code_value)
        current = current.parent
    
    full_path = ' → '.join(path_parts)
    
    return jsonify({
        'success': True,
        'code': {
            'id': code.id,
            'dictionary_id': code.dictionary_id,
            'dictionary_name': code.dictionary.dict_name_ar or code.dictionary.dict_name,
            'code_value': code.code_value,
            'code_value_ar': code.code_value_ar,
            'description': code.code_description,
            'display_color': code.display_color,
            'display_sequence': code.display_sequence,
            'level': code.level,
            'full_path': code.full_path,
            'full_path_display': full_path,
            'is_active': code.is_active,
            'parent_id': code.parent_id,
            'parent_value': code.parent.code_value if code.parent else None,
            'assignments_count': code.assignments.count(),
            'created_at': code.created_at.strftime('%Y-%m-%d %H:%M') if code.created_at else None,
            'created_by': code.creator.full_name if code.creator else None
        }
    })
# ============================================
# API Routes للإعدادات (Defaults & Settings)
# ============================================

@project_bp.route('/api/<int:project_id>/update-defaults', methods=['POST'])
@login_required
def api_update_defaults(project_id):
    """تحديث القيم الافتراضية"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        if not project.defaults:
            project.defaults = {}
        
        project.defaults.update(data)
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/<int:project_id>/update-settings', methods=['POST'])
@login_required
def api_update_settings(project_id):
    """تحديث إعدادات المشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        if 'last_summarized' in data and data['last_summarized']:
            project.last_summarized = datetime.strptime(data['last_summarized'], '%Y-%m-%d')
        if 'summarize_level' in data:
            project.summarize_level = int(data['summarize_level'])
        if 'fiscal_year_start' in data:
            project.fiscal_year_start = data['fiscal_year_start']
        if 'wbs_separator' in data:
            project.wbs_separator = data['wbs_separator']
        if 'float_threshold' in data:
            project.float_threshold = float(data['float_threshold'])
        
        if not project.settings:
            project.settings = {}
        
        project.settings['baseline_type'] = data.get('baseline_type')
        project.settings['critical_definition'] = data.get('critical_definition')
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/<int:project_id>/update-calculation-settings', methods=['POST'])
@login_required
def api_update_calculation_settings(project_id):
    """تحديث إعدادات الحسابات"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        if not project.calculation_settings:
            project.calculation_settings = {}
        
        project.calculation_settings.update(data)
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# API Routes للحسابات (Calculations)
# ============================================

@project_bp.route('/api/<int:project_id>/run-schedule', methods=['POST'])
@login_required
def api_run_schedule(project_id):
    """تشغيل الجدولة"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    try:
        activities = Activity.query.filter_by(project_id=project_id).all()
        
        # تحديث مؤشرات الأداء
        if project.performance and activities:
            total_planned = sum(a.planned_value or 0 for a in activities)
            total_earned = sum(a.earned_value or 0 for a in activities)
            total_actual = sum(a.actual_cost or 0 for a in activities)
            
            if total_planned > 0:
                project.performance.spi = total_earned / total_planned
            if total_actual > 0:
                project.performance.cpi = total_earned / total_actual
            
            project.performance.csi = (project.performance.spi or 1) * (project.performance.cpi or 1)
            
            # تحديث EAC
            if project.performance.cpi and project.performance.cpi > 0:
                project.performance.eac = total_actual + (total_planned - total_earned) / project.performance.cpi
                project.performance.etc = project.performance.eac - total_actual
                project.performance.vac = total_planned - project.performance.eac
        
        # تحديث التقدم
        if project.progress and activities:
            total_weight = sum(a.weight for a in activities)
            if total_weight > 0:
                weighted_progress = sum(a.progress_percentage * a.weight for a in activities)
                project.progress.progress_percentage = weighted_progress / total_weight
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Schedule completed successfully',
            'performance': {
                'spi': project.performance.spi if project.performance else 1,
                'cpi': project.performance.cpi if project.performance else 1,
                'eac': project.performance.eac if project.performance else 0
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# API Routes للملخصات (Summaries)
# ============================================

@project_bp.route('/api/<int:project_id>/budget-summary')
@login_required
def api_budget_summary(project_id):
    """الحصول على ملخص الميزانية"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    # حساب الميزانية الموزعة
    activities = Activity.query.filter_by(project_id=project_id).all()
    distributed_budget = sum(a.planned_value or 0 for a in activities)
    
    current_budget = project.budget.current_budget if project.budget else 0
    unallocated_budget = max(0, current_budget - distributed_budget)
    
    # حساب خطة الصرف
    spending_items = SpendingPlanItem.query.filter_by(project_id=project_id).all()
    total_spending = sum(i.planned_amount for i in spending_items)
    total_benefit = sum(i.benefit_amount for i in spending_items)
    
    # حساب الفروق
    if spending_items:
        last_item = spending_items[-1]
        undistributed_variance = last_item.undistributed_variance if hasattr(last_item, 'undistributed_variance') else 0
        total_benefit_tally = last_item.benefit_tally if hasattr(last_item, 'benefit_tally') else total_benefit
    else:
        undistributed_variance = 0
        total_benefit_tally = 0
    
    return jsonify({
        'success': True,
        'summary': {
            'current_budget': current_budget,
            'unallocated_budget': unallocated_budget,
            'distributed_budget': distributed_budget,
            'current_variance': current_budget - (project.cost.total_actual_cost if project.cost else 0),
            'total_spending_plan': total_spending,
            'total_benefit_plan': total_benefit,
            'undistributed_variance': undistributed_variance,
            'total_spending_tally': total_spending,
            'total_benefit_tally': total_benefit_tally
        }
    })


# ============================================
# API Routes للمصادر (Sources)
# ============================================

@project_bp.route('/api/eps/<int:eps_id>')
@login_required
def api_eps_detail(eps_id):
    """الحصول على تفاصيل EPS"""
    eps = EPS.query.get_or_404(eps_id)
    
    if eps.org_id != current_user.org_id and current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    return jsonify({
        'success': True,
        'eps': {
            'id': eps.id,
            'name': eps.name,
            'code': eps.eps_code
        }
    })


@project_bp.route('/api/funding/<int:source_id>/delete', methods=['POST'])
@login_required
def api_delete_funding(source_id):
    """حذف مصدر تمويل"""
    source = FundingSource.query.get_or_404(source_id)
    
    project = Project.query.get(source.project_id)
    if project.created_by != current_user.id and current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    try:
        db.session.delete(source)
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    
@project_bp.route('/api/<int:project_id>/resource-assignments')
@login_required
def api_resource_assignments(project_id):
    """API لجلب تعيينات الموارد للمشروع"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    # Get all resource assignments for the project's activities
    assignments = ActivityResource.query.join(Activity).filter(Activity.project_id == project_id).all()
    
    result = []
    total_cost = 0
    total_units = 0
    
    for ass in assignments:
        resource = ass.resource
        activity = ass.activity
        
        if not resource or not activity:
            continue
        
        planned_cost = (ass.planned_quantity or 0) * (resource.cost_per_unit or 0)
        actual_cost = (ass.actual_quantity or 0) * (resource.cost_per_unit or 0)
        remaining = (ass.planned_quantity or 0) - (ass.actual_quantity or 0)
        
        total_cost += actual_cost
        total_units += ass.actual_quantity or 0
        
        result.append({
            'id': ass.id,
            'resource_id': resource.resource_id,
            'resource_name': resource.name,
            'resource_type': resource.resource_type,
            'activity_id': activity.id,
            'activity_code': activity.activity_code,
            'planned_units': ass.planned_quantity,
            'actual_units': ass.actual_quantity,
            'remaining_units': max(0, remaining),
            'planned_cost': planned_cost,
            'actual_cost': actual_cost,
            'rate_type': ass.rate_type if hasattr(ass, 'rate_type') else 'Standard',
            'drive_dates': ass.drive_dates if hasattr(ass, 'drive_dates') else False
        })
    
    # Resource statistics
    stats = {
        'labor_count': sum(1 for a in result if a['resource_type'] == 'labor'),
        'equipment_count': sum(1 for a in result if a['resource_type'] == 'equipment'),
        'material_count': sum(1 for a in result if a['resource_type'] == 'material'),
        'total_cost': total_cost,
        'total_units': total_units
    }
    
    # Resource chart data
    resources = Resource.query.filter_by(org_id=project.org_id).all()
    chart_data = []
    for r in resources:
        assigned = sum(a.planned_quantity or 0 for a in ActivityResource.query.filter_by(resource_id=r.id)
                      .join(Activity).filter(Activity.project_id == project_id).all())
        chart_data.append({
            'name': r.name,
            'available': r.available_quantity or 0,
            'assigned': assigned
        })
    
    return jsonify({
        'success': True,
        'assignments': result,
        'stats': stats,
        'chart_data': chart_data
    })


@project_bp.route('/api/<int:project_id>/resource-assign', methods=['POST'])
@login_required
def api_resource_assign(project_id):
    """API لتعيين مورد لنشاط"""
    project = Project.query.get_or_404(project_id)
    
    if project.org_id != current_user.org_id:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    resource_id = data.get('resource_id')
    activity_id = data.get('activity_id')
    planned_units = data.get('planned_units', 1)
    rate_type = data.get('rate_type', 'Standard')
    drive_dates = data.get('drive_dates', True)
    start_date = data.get('start_date')
    finish_date = data.get('finish_date')
    
    if not resource_id or not activity_id:
        return jsonify({'success': False, 'error': 'المورد والنشاط مطلوبان'}), 400
    
    resource = Resource.query.get_or_404(resource_id)
    activity = Activity.query.get_or_404(activity_id)
    
    # التحقق من الكمية المتاحة
    # if planned_units > resource.available_quantity:
    #     return jsonify({
    #         'success': False,
    #         'error': f'الكمية غير متوفرة. المتاح: {resource.available_quantity} {resource.unit}'
    #     }), 400
    
    try:
        # إنشاء تخصيص جديد
        assignment = ActivityResource(
            activity_id=activity_id,
            resource_id=resource_id,
            planned_quantity=planned_units,
            actual_quantity=0,
            remaining_quantity=planned_units,
            planned_cost=planned_units * resource.cost_per_unit,
            rate_type=rate_type,
            allocated=drive_dates,
            created_by=current_user.id
        )
        
        # تحديث التواريخ إذا كانت محددة
        if start_date:
            assignment.start_date = datetime.strptime(start_date, '%Y-%m-%d')
        if finish_date:
            assignment.finish_date = datetime.strptime(finish_date, '%Y-%m-%d')
        
        db.session.add(assignment)
        
        # تحديث الكمية المخصصة في المورد
        # resource.udf_values['total_allocated'] = (resource.udf_values['total_allocated'] or 0) + planned_units
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'assignment': {
                'id': assignment.id,
                'resource_id': resource.resource_id,
                'resource_name': resource.name,
                'activity_name': activity.activity_name,
                'planned_units': planned_units,
                'planned_cost': planned_units * resource.cost_per_unit
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/resource-assignment/<int:assignment_id>/update', methods=['POST'])
@login_required
def api_resource_assignment_update(assignment_id):
    """API لتحديث تعيين مورد"""
    assignment = ActivityResource.query.get_or_404(assignment_id)
    resource = assignment.resource
    
    # التحقق من الصلاحية
    activity = assignment.activity
    project = activity.project
    if project.org_id != current_user.org_id:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        if 'planned_units' in data:
            new_units = float(data['planned_units'])
            old_units = assignment.planned_quantity
            
            # تحديث الكمية المخصصة في المورد
            diff = new_units - old_units
            resource.total_allocated = (resource.total_allocated or 0) + diff
            
            assignment.planned_quantity = new_units
            assignment.remaining_quantity = new_units - assignment.actual_quantity
            assignment.planned_cost = new_units * resource.cost_per_unit
            
        if 'actual_units' in data:
            new_actual = float(data['actual_units'])
            assignment.actual_quantity = new_actual
            assignment.remaining_quantity = assignment.planned_quantity - new_actual
            assignment.actual_cost = new_actual * resource.cost_per_unit
            
        if 'rate_type' in data:
            assignment.rate_type = data['rate_type']
            assignment.planned_cost = assignment.planned_quantity * resource.cost_per_unit
            
        if 'drive_dates' in data:
            assignment.drive_dates = data['drive_dates']
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'planned_cost': assignment.planned_cost,
            'actual_cost': assignment.actual_cost,
            'remaining_units': assignment.remaining_quantity
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/resource-assignment/<int:assignment_id>/delete', methods=['POST'])
@login_required
def api_resource_assignment_delete(assignment_id):
    """API لحذف تعيين مورد"""
    assignment = ActivityResource.query.get_or_404(assignment_id)
    
    activity = assignment.activity
    project = activity.project
    if project.org_id != current_user.org_id:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    try:
        # تحرير الكمية المخصصة من المورد
        resource = assignment.resource
        resource.total_allocated = (resource.total_allocated or 0) - assignment.planned_quantity
        
        db.session.delete(assignment)
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/<int:project_id>/update-resource-defaults', methods=['POST'])
@login_required
def api_update_resource_defaults(project_id):
    """API لتحديث إعدادات الموارد الافتراضية للمشروع"""
    project = Project.query.get_or_404(project_id)
    
    if project.org_id != current_user.org_id:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        if not hasattr(project, 'resource_defaults') or not project.resource_defaults:
            project.resource_defaults = {}
        
        project.resource_defaults['rate_type'] = data.get('rate_type', 'Standard Rate')
        project.resource_defaults['drive_activity_dates'] = data.get('drive_activity_dates', True)
        project.resource_defaults['allow_multiple_assignments'] = data.get('allow_multiple_assignments', True)
        
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/<int:project_id>/resource-chart-data')
@login_required
def api_resource_chart_data(project_id):
    """API لجلب بيانات الرسم البياني للموارد"""
    project = Project.query.get_or_404(project_id)
    
    if project.org_id != current_user.org_id:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    activities = Activity.query.filter_by(project_id=project_id).all()
    resources = Resource.query.filter_by(org_id=current_user.org_id, is_active=True).limit(10).all()
    
    labels = []
    available = []
    assigned = []
    
    for resource in resources:
        labels.append(resource.name[:20])
        available.append(resource.available_quantity)
        
        total_assigned = 0
        for activity in activities:
            for assignment in activity.resources:
                if assignment.resource_id == resource.id:
                    total_assigned += assignment.planned_quantity
        
        assigned.append(total_assigned)
    
    return jsonify({
        'success': True,
        'labels': labels,
        'available': available,
        'assigned': assigned
    })


@project_bp.route('/api/<int:project_id>/import-resources', methods=['POST'])
@login_required
def api_import_resources(project_id):
    """API لاستيراد موارد من Excel"""
    project = Project.query.get_or_404(project_id)
    
    if project.org_id != current_user.org_id:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'لم يتم رفع ملف'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'الملف فارغ'}), 400
    
    try:
        # قراءة الملف
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        results = {'success': [], 'failed': []}
        
        for _, row in df.iterrows():
            try:
                # البحث عن المورد
                resource_code = row.get('Resource Code') or row.get('الكود')
                resource_name = row.get('Resource Name') or row.get('الاسم')
                
                resource = None
                if resource_code:
                    resource = Resource.query.filter_by(
                        org_id=current_user.org_id,
                        resource_id=resource_code
                    ).first()
                
                if not resource and resource_name:
                    resource = Resource.query.filter_by(
                        org_id=current_user.org_id,
                        name=resource_name
                    ).first()
                
                if not resource:
                    results['failed'].append({
                        'name': resource_name or resource_code,
                        'error': 'المورد غير موجود'
                    })
                    continue
                
                # البحث عن النشاط
                activity_code = row.get('Activity Code') or row.get('كود النشاط')
                activity = None
                if activity_code:
                    activity = Activity.query.filter_by(
                        project_id=project_id,
                        activity_id=activity_code
                    ).first()
                
                if not activity:
                    results['failed'].append({
                        'name': resource.name,
                        'error': f'النشاط {activity_code} غير موجود'
                    })
                    continue
                
                quantity = float(row.get('Quantity') or row.get('الكمية', 0))
                
                if quantity <= 0:
                    results['failed'].append({
                        'name': resource.name,
                        'error': 'الكمية يجب أن تكون أكبر من صفر'
                    })
                    continue
                
                # إنشاء التخصيص
                assignment = ActivityResource(
                    activity_id=activity.id,
                    resource_id=resource.id,
                    planned_quantity=quantity,
                    actual_quantity=0,
                    remaining_quantity=quantity,
                    allocated_quantity=quantity,
                    planned_cost=quantity * resource.cost_per_unit,
                    created_by=current_user.id
                )
                db.session.add(assignment)
                
                resource.total_allocated = (resource.total_allocated or 0) + quantity
                
                results['success'].append({
                    'name': resource.name,
                    'activity': activity.activity_name,
                    'quantity': quantity
                })
                
            except Exception as e:
                results['failed'].append({
                    'name': row.get('Resource Name', 'غير معروف'),
                    'error': str(e)
                })
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'results': results,
            'summary': {
                'success': len(results['success']),
                'failed': len(results['failed'])
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/<int:project_id>/export-resources')
@login_required
def api_export_resources(project_id):
    """API لتصدير موارد المشروع إلى Excel"""
    project = Project.query.get_or_404(project_id)
    
    if project.org_id != current_user.org_id:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    activities = Activity.query.filter_by(project_id=project_id).all()
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Resource Assignments"
    
    # تحديد الأعمدة
    headers = ['Resource Code', 'Resource Name', 'Type', 'Activity Code', 'Activity Name', 
               'Planned Units', 'Actual Units', 'Remaining Units', 'Planned Cost', 'Actual Cost', 'Rate Type']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # إضافة البيانات
    row_num = 2
    for activity in activities:
        for assignment in activity.resources:
            resource = assignment.resource
            
            ws.cell(row=row_num, column=1, value=resource.resource_id)
            ws.cell(row=row_num, column=2, value=resource.name)
            ws.cell(row=row_num, column=3, value=resource.resource_type)
            ws.cell(row=row_num, column=4, value=activity.activity_id)
            ws.cell(row=row_num, column=5, value=activity.activity_name)
            ws.cell(row=row_num, column=6, value=assignment.planned_quantity)
            ws.cell(row=row_num, column=7, value=assignment.actual_quantity)
            ws.cell(row=row_num, column=8, value=assignment.remaining_quantity)
            ws.cell(row=row_num, column=9, value=assignment.planned_cost)
            ws.cell(row=row_num, column=10, value=assignment.actual_cost)
            ws.cell(row=row_num, column=11, value=getattr(assignment, 'rate_type', 'Standard'))
            
            row_num += 1
    
    # ضبط عرض الأعمدة
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'project_{project.project_code}_resources.xlsx'
    )


@project_bp.route('/api/<int:project_id>/resource-template')
@login_required
def api_resource_template(project_id):
    """تحميل قالب استيراد الموارد"""
    project = Project.query.get_or_404(project_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resource Import Template"
    
    headers = ['Resource Code', 'Resource Name', 'Activity Code', 'Quantity', 'Rate Type', 'Notes']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # إضافة أمثلة
    examples = [
        ['RES-001', 'حفار كوماتسو', 'A1000', 2, 'Standard', 'مطلوب للتسليم'],
        ['RES-002', 'حديد تسليح', 'A1010', 50, 'Standard', ''],
        ['', 'عامل بناء', 'A1020', 10, 'Overtime', '']
    ]
    
    for row, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            ws.cell(row=row, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'project_{project.project_code}_resource_template.xlsx'
    )

@project_bp.route('/api/resources/list')
@login_required
def api_resources_list():
    """API لقائمة الموارد للمؤسسة"""
    org_id = current_user.org_id
    
    resources = Resource.query.filter_by(org_id=org_id).all()
    
    return jsonify({
        'success': True,
        'resources': [{
            'id': r.id,
            'resource_id': r.resource_id,
            'name': r.name,
            'type': r.resource_type,
            'unit': r.unit,
            'cost_per_unit': r.cost_per_unit,
            'available_quantity': r.available_quantity
        } for r in resources]
    })

# ============================================
# API Routes للحقول المخصصة (UDF)
# ============================================

@project_bp.route('/api/udf/<int:udf_id>')
@login_required
def api_udf_detail(udf_id):
    """API لجلب تفاصيل حقل مخصص"""
    udf = ProjectUDF.query.get_or_404(udf_id)
    
    project = Project.query.get(udf.project_id)
    if project.created_by != current_user.id and current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    return jsonify({
        'success': True,
        'udf': {
            'id': udf.id,
            'udf_name': udf.udf_name,
            'udf_value': udf.udf_value,
            'udf_type': udf.udf_type
        }
    })


@project_bp.route('/api/<int:project_id>/udf/<int:udf_id>/update', methods=['POST'])
@login_required
def api_update_udf(project_id, udf_id):
    """API لتحديث حقل مخصص"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    udf = ProjectUDF.query.get_or_404(udf_id)
    data = request.get_json()
    
    try:
        udf.udf_name = data.get('udf_name', udf.udf_name)
        udf.udf_value = data.get('udf_value')
        udf.udf_type = data.get('udf_type', udf.udf_type)
        udf.updated_at = datetime.utcnow()
        udf.updated_by = current_user.id
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/udf/<int:udf_id>/delete', methods=['POST'])
@login_required
def api_delete_udf(udf_id):
    """API لحذف حقل مخصص"""
    udf = ProjectUDF.query.get_or_404(udf_id)
    
    project = Project.query.get(udf.project_id)
    if project.created_by != current_user.id and current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    try:
        db.session.delete(udf)
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# API Routes للمخاطر (Risks)
# ============================================

@project_bp.route('/api/<int:project_id>/risk', methods=['POST'])
@login_required
def api_add_risk(project_id):
    """API لإضافة خطر جديد"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        # Create risk code if not provided
        risk_code = data.get('risk_code')
        if not risk_code:
            last_risk = Risk.query.filter_by(project_id=project_id).order_by(Risk.id.desc()).first()
            if last_risk and last_risk.risk_code:
                last_num = int(last_risk.risk_code.split('-')[-1]) if '-' in last_risk.risk_code else 1
                risk_code = f"RISK-{last_num + 1:04d}"
            else:
                risk_code = "RISK-0001"
        
        # Calculate severity
        probability = int(data.get('probability', 50))
        impact_map = {'very_low': 20, 'low': 40, 'medium': 60, 'high': 80, 'very_high': 100}
        impact = impact_map.get(data.get('impact'), 60)
        severity = (probability * impact) / 100
        
        if severity >= 70:
            risk_level = 'high'
        elif severity >= 40:
            risk_level = 'medium'
        else:
            risk_level = 'low'
        
        risk = Risk(
            project_id=project_id,
            risk_code=risk_code,
            title=data.get('title'),
            description=data.get('description'),
            category=data.get('category'),
            probability=probability,
            impact=data.get('impact'),
            severity=severity,
            risk_level=risk_level,
            status=data.get('status', 'identified'),
            mitigation_plan=data.get('mitigation_plan'),
            contingency_plan=data.get('contingency_plan'),
            target_mitigation_date=datetime.strptime(data.get('target_date'), '%Y-%m-%d').date() if data.get('target_date') else None,
            owner_id=data.get('owner_id'),
            created_by=current_user.id
        )
        
        db.session.add(risk)
        db.session.commit()
        
        return jsonify({'success': True, 'risk_id': risk.id})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/risk/<int:risk_id>')
@login_required
def api_risk_detail(risk_id):
    """API لجلب تفاصيل خطر"""
    risk = Risk.query.get_or_404(risk_id)
    
    project = Project.query.get(risk.project_id)
    if project.created_by != current_user.id and current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    return jsonify({
        'success': True,
        'risk': {
            'id': risk.id,
            'risk_code': risk.risk_code,
            'title': risk.title,
            'description': risk.description,
            'category': risk.category,
            'probability': risk.probability,
            'impact': risk.impact,
            'severity': risk.severity,
            'risk_level': risk.risk_level,
            'status': risk.status,
            'mitigation_plan': risk.mitigation_plan,
            'contingency_plan': risk.contingency_plan,
            'target_mitigation_date': risk.target_mitigation_date.isoformat() if risk.target_mitigation_date else None,
            'owner_id': risk.owner_id
        }
    })


@project_bp.route('/api/risk/<int:risk_id>/update', methods=['POST'])
@login_required
def api_update_risk(risk_id):
    """API لتحديث خطر"""
    risk = Risk.query.get_or_404(risk_id)
    
    project = Project.query.get(risk.project_id)
    if project.created_by != current_user.id and current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    
    try:
        risk.title = data.get('title', risk.title)
        risk.description = data.get('description', risk.description)
        risk.category = data.get('category', risk.category)
        risk.probability = int(data.get('probability', risk.probability))
        risk.impact = data.get('impact', risk.impact)
        risk.status = data.get('status', risk.status)
        risk.mitigation_plan = data.get('mitigation_plan', risk.mitigation_plan)
        risk.contingency_plan = data.get('contingency_plan', risk.contingency_plan)
        risk.owner_id = data.get('owner_id', risk.owner_id)
        
        if data.get('target_date'):
            risk.target_mitigation_date = datetime.strptime(data['target_date'], '%Y-%m-%d').date()
        
        # Recalculate severity
        impact_map = {'very_low': 20, 'low': 40, 'medium': 60, 'high': 80, 'very_high': 100}
        impact = impact_map.get(risk.impact, 60)
        risk.severity = (risk.probability * impact) / 100
        
        if risk.severity >= 70:
            risk.risk_level = 'high'
        elif risk.severity >= 40:
            risk.risk_level = 'medium'
        else:
            risk.risk_level = 'low'
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/risk/<int:risk_id>/delete', methods=['POST'])
@login_required
def api_delete_risk(risk_id):
    """API لحذف خطر"""
    risk = Risk.query.get_or_404(risk_id)
    
    project = Project.query.get(risk.project_id)
    if project.created_by != current_user.id and current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    try:
        db.session.delete(risk)
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# API Routes للمستندات (Documents)
# ============================================

@project_bp.route('/api/<int:project_id>/document/upload', methods=['POST'])
@login_required
def api_upload_document(project_id):
    """API لرفع مستند"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'لم يتم رفع ملف'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'لم يتم اختيار ملف'}), 400
    
    try:
        # Create upload directory
        upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'project_documents', str(project_id))
        os.makedirs(upload_folder, exist_ok=True)
        
        # Secure filename and save
        filename = secure_filename(file.filename)
        file_path = os.path.join(upload_folder, f"{uuid.uuid4()}_{filename}")
        file.save(file_path)
        
        # Determine file type
        ext = filename.split('.')[-1].lower() if '.' in filename else ''
        file_type_map = {
            'pdf': 'pdf',
            'doc': 'doc', 'docx': 'doc',
            'xls': 'xls', 'xlsx': 'xls',
            'jpg': 'jpg', 'jpeg': 'jpg', 'png': 'png', 'gif': 'gif'
        }
        file_type = file_type_map.get(ext, 'other')
        
        # Create document record
        document = ProjectDocument(
            project_id=project_id,
            filename=filename,
            original_filename=file.filename,
            title=request.form.get('title', filename),
            description=request.form.get('description'),
            category=request.form.get('category', 'other'),
            file_type=file_type,
            file_url=f"/static/uploads/project_documents/{project_id}/{os.path.basename(file_path)}",
            file_size=os.path.getsize(file_path),
            requires_approval=request.form.get('requires_approval') == 'on',
            uploaded_by=current_user.id,
            uuid=str(uuid.uuid4())
        )
        
        db.session.add(document)
        db.session.commit()
        
        return jsonify({'success': True, 'document_id': document.id})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/document/<int:document_id>/delete', methods=['POST'])
@login_required
def api_delete_document(document_id):
    """API لحذف مستند"""
    document = ProjectDocument.query.get_or_404(document_id)
    
    project = Project.query.get(document.project_id)
    if project.created_by != current_user.id and current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    try:
        # Delete file from filesystem
        if document.file_url:
            file_path = os.path.join(current_app.root_path, 'static', document.file_url.replace('/static/', ''))
            if os.path.exists(file_path):
                os.remove(file_path)
        
        db.session.delete(document)
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/<int:project_id>/documents/stats')
@login_required
def api_document_stats(project_id):
    """API لإحصائيات المستندات"""
    project = check_project_access(project_id)
    if not project:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    documents = ProjectDocument.query.filter_by(project_id=project_id).all()
    
    stats = {
        'pdf_count': sum(1 for d in documents if d.file_type == 'pdf'),
        'doc_count': sum(1 for d in documents if d.file_type == 'doc'),
        'excel_count': sum(1 for d in documents if d.file_type == 'xls'),
        'image_count': sum(1 for d in documents if d.file_type in ['jpg', 'png', 'gif']),
        'total_count': len(documents),
        'total_size': sum(d.file_size or 0 for d in documents)
    }
    
    return jsonify({'success': True, 'stats': stats})

# app/routes/project_routes.py

# from app.models import ResourceRequest, ResourceRequestItem, ResourceRequestNotification
# from app.services.resource_request_service import ResourceRequestService
# from datetime import datetime, timedelta

@project_bp.route('/<int:project_id>/resource-requests')
@login_required
def project_resource_requests(project_id):
    """عرض طلبات التوريد لمشروع معين"""
    project = Project.query.get_or_404(project_id)
    
    if project.org_id != current_user.org_id:
        flash('غير مصرح بالوصول', 'danger')
        return redirect(url_for('company.dashboard'))
    
    # جلب جميع طلبات التوريد للمشروع
    requests = ResourceRequest.query.filter_by(
        project_id=project_id
    ).order_by(ResourceRequest.created_at.desc()).all()
    
    return render_template('projects/resource_requests.html',
                         project=project,
                         requests=requests,
                         now=datetime.now())


# app/routes/project_routes.py

# from app.models.resource_models import ResourceRequest, ResourceRequestItem
# from app.models.primavera_models import Resource
# from datetime import datetime, timedelta

@project_bp.route('/<int:project_id>/resource-request/create', methods=['GET', 'POST'])
@login_required
def create_resource_request(project_id):
    """إنشاء طلب توريد جديد"""
    project = Project.query.get_or_404(project_id)
    
    if project.org_id != current_user.org_id:
        flash('غير مصرح بالوصول', 'danger')
        return redirect(url_for('company.dashboard'))
    
    # ✅ جلب المواد التي لم يتم طلبها بالكامل بعد
    # المواد من نوع material في المشروع
    project_materials = []
    
    # جلب جميع المواد المستخدمة في المشروع
    activities = Activity.query.filter_by(project_id=project_id).all()
    material_ids = set()
    
    for activity in activities:
        for assignment in activity.resources:
            if assignment.resource.resource_type == 'material' or assignment.resource.resource_type == 'equipment':
                material_ids.add(assignment.resource_id)
    
    # المواد التي تم طلبها بالفعل
    existing_requests = ResourceRequest.query.filter_by(project_id=project_id).all()
    requested_material_ids = set()
    
    for req in existing_requests:
        for item in req.items:
            if item.resource_id:
                requested_material_ids.add(item.resource_id)
    
    # المواد المطلوب عرضها (الموجودة في المشروع ولم يتم طلبها بالكامل)
    for material_id in material_ids:
        if material_id not in requested_material_ids:
            material = Resource.query.get(material_id)
            if material:
                project_materials.append(material)
    
    # ✅ جلب جميع الموردين من المستخدمين
    from app.models import User
    suppliers = User.query.filter_by(
        org_id=project.org_id,
        role='supplier',
        is_user_active=True
    ).all()
    
    # حساب التاريخ الافتراضي
    today = datetime.now().date()
    default_date = today + timedelta(days=7)
    default_date_str = default_date.strftime('%Y-%m-%d')
    
    if request.method == 'POST':
        supplier_id = request.form.get('supplier_id')
        required_date = datetime.strptime(request.form.get('required_date'), '%Y-%m-%d').date()
        site_location = request.form.get('site_location')
        notes = request.form.get('notes')
        
        # جمع المواد المختارة
        material_ids_selected = request.form.getlist('material_ids[]')
        quantities = request.form.getlist('quantities[]')
        
        if not material_ids_selected:
            flash('الرجاء اختيار مادة واحدة على الأقل', 'danger')
            return redirect(url_for('projects.create_resource_request', project_id=project_id))
        
        try:
            # إنشاء طلب التوريد
            request_obj = ResourceRequest(
                org_id=project.org_id,
                project_id=project_id,
                supplier_id=supplier_id,
                resources=[],  # سيتم تعبئته لاحقاً
                required_date=required_date,
                status='pending',
                site_location=site_location or project.location.site_name,
                notes=notes,
                created_by=current_user.id
            )
            db.session.add(request_obj)
            db.session.flush()
            
            # إضافة بنود الطلب
            resources_list = []
            for i in range(len(material_ids_selected)):
                material_id = int(material_ids_selected[i])
                quantity = float(quantities[i]) if quantities[i] else 0
                
                if quantity <= 0:
                    continue
                
                material = Resource.query.get(material_id)
                if material:
                    request_item = ResourceRequestItem(
                        request_id=request_obj.id,
                        resource_id=material.id,
                        resource_name=material.name,
                        unit=material.unit,
                        required_quantity=quantity,
                        remaining_quantity=quantity
                    )
                    db.session.add(request_item)
                    
                    resources_list.append({
                        'id': material.id,
                        'name': material.name,
                        'quantity': quantity,
                        'unit': material.unit
                    })
            
            # تحديث حقل resources في الطلب
            request_obj.resources = resources_list
            
            db.session.commit()
            
            NotificationService.resource_request_created(request_obj)
            
            flash('تم إنشاء طلب التوريد بنجاح', 'success')
            return redirect(url_for('projects.project_resource_requests', project_id=project_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    return render_template('projects/create_resource_request.html',
                         project=project,
                         materials=project_materials,
                         suppliers=suppliers,
                         default_date_str=default_date_str,
                         now=datetime.now())


@project_bp.route('/api/resource-request/<int:request_id>/update-status', methods=['POST'])
@login_required
def update_resource_request_status(request_id):
    """تحديث حالة طلب التوريد"""
    request_obj = ResourceRequest.query.get_or_404(request_id)
    project = request_obj.project
    
    if project.org_id != current_user.org_id:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    data = request.get_json()
    new_status = data.get('status')
    notes = data.get('notes')
    
    if new_status not in ['pending', 'started', 'completed', 'cancelled']:
        return jsonify({'success': False, 'error': 'حالة غير صالحة'}), 400
    
    try:
        old_status = request_obj.status
        request_obj.status = new_status
        request_obj.updated_at = datetime.utcnow()
        
        if new_status == 'started' and not request_obj.started_at:
            request_obj.started_at = datetime.utcnow()
        elif new_status == 'completed':
            request_obj.completed_at = datetime.utcnow()
        
        # تسجيل التحديث
        update = ResourceRequestUpdate(
            request_id=request_id,
            old_status=old_status,
            new_status=new_status,
            message=notes,
            updated_by=current_user.id
        )
        db.session.add(update)
        db.session.commit()
        
        # إرسال إشعار للمورد
        if new_status in ['started', 'completed', 'cancelled']:
            NotificationService.resource_request_updated(request_obj,new_status)
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@project_bp.route('/api/resource-request/<int:request_id>/remind', methods=['POST'])
@login_required
def remind_supplier(request_id):
    """إرسال تذكير للمورد"""
    request_obj = ResourceRequest.query.get_or_404(request_id)
    project = request_obj.project
    
    if project.org_id != current_user.org_id:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    
    try:

        NotificationService.remind_supplier_notification(request_obj)
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
@project_bp.route('/delivery/<int:delivery_id>/confirm', methods=['GET', 'POST'])
@login_required
def confirm_delivery(delivery_id):
    """تأكيد أو رفض تسليم المواد"""
    delivery = ResourceDelivery.query.get_or_404(delivery_id)
    resource_request = delivery.request
    project = resource_request.project
    
    # التحقق من الصلاحية
    if project.project_manager_id != current_user.id and not current_user.is_admin():
        flash('غير مصرح لك', 'danger')
        return redirect(url_for('company.dashboard'))
    
    # التحقق من حالة التسليم
    # if delivery.status != 'pending':
    #     flash('تم معالجة هذا التسليم مسبقاً', 'warning')
    #     return redirect(url_for('projects.project_resource_requests', project_id=project.id))
    
    # جلب عناصر الطلب
    request_items = ResourceRequestItem.query.filter_by(
        request_id=resource_request.id
    ).all()
    
    if request.method == 'POST':
        data = request.get_json()
        action = data.get('action')
        notes = data.get('notes', '')
        
        try:
            if action == 'confirm':
                # تأكيد التسليم
                delivery.status = 'confirmed'
                delivery.confirmed_by = current_user.id
                delivery.confirmed_at = datetime.utcnow()
                delivery.confirmation_notes = notes
                
                # تحديث حالة الطلب إذا لزم الأمر
                if resource_request.status == 'partially_delivered':
                    all_completed = all(item.is_completed for item in request_items)
                    if all_completed:
                        resource_request.status = 'completed'
                        resource_request.completed_at = datetime.utcnow()
                
                db.session.commit()
                # ✅ تحديث المؤشرات
                UpdateService.update_resource_request_metrics(resource_request.id)
                # إرسال إشعار للمورد بتأكيد الاستلام
                NotificationService.delivery_confirmed(delivery, resource_request, True, notes)
                
                return jsonify({
                    'success': True,
                    'message': 'تم تأكيد استلام المواد بنجاح'
                })
                
            elif action == 'reject':
                # رفض التسليم
                if not notes:
                    return jsonify({
                        'success': False,
                        'error': 'الرجاء إدخال سبب الرفض'
                    }), 400
                
                delivery.status = 'rejected'
                delivery.rejection_reason = notes
                delivery.confirmed_by = current_user.id
                delivery.confirmed_at = datetime.utcnow()
                
                # إرجاع الكميات إلى العناصر
                for delivered in delivery.delivered_items:
                    item = ResourceRequestItem.query.get(delivered['item_id'])
                    if item:
                        item.delivered_quantity -= delivered['quantity']
                        item.remaining_quantity = item.required_quantity - item.delivered_quantity
                        item.is_completed = False
                        item.updated_at = datetime.utcnow()
                
                # إعادة حالة الطلب إذا لزم الأمر
                if resource_request.status == 'completed':
                    resource_request.status = 'partially_delivered'
                elif resource_request.status == 'partially_delivered':
                    # تحقق إذا كان هناك أي تسليمات أخرى
                    other_deliveries = ResourceDelivery.query.filter(
                        ResourceDelivery.request_id == resource_request.id,
                        ResourceDelivery.id != delivery_id,
                        ResourceDelivery.status == 'confirmed'
                    ).count()
                    
                    if other_deliveries == 0:
                        resource_request.status = 'started'
                
                resource_request.updated_at = datetime.utcnow()
                
                db.session.commit()
                
                # إرسال إشعار للمورد بالرفض
                NotificationService.delivery_confirmed(delivery, resource_request, False, notes)
                
                return jsonify({
                    'success': True,
                    'message': 'تم رفض التسليم بنجاح'
                })
            
            else:
                return jsonify({
                    'success': False,
                    'error': 'إجراء غير صالح'
                }), 400
                
        except Exception as e:
            db.session.rollback()
            return jsonify({
                'success': False,
                'error': str(e)
            }), 500
    
    # GET request - عرض صفحة التأكيد
    return render_template(
        'projects/confirm_delivery.html',
        delivery=delivery,
        resource_request=resource_request,
        request_items=request_items,
        project=project
    )

# أضف هذه المسارات في project_routes.py

@project_bp.route('/<int:project_id>/equipment-requests')
@login_required
def project_equipment_requests(project_id):
    """عرض طلبات المعدات لمشروع معين"""
    from app.models import EquipmentRequest
    
    project = Project.query.get_or_404(project_id)
    
    if project.org_id != current_user.org_id:
        flash('غير مصرح بالوصول', 'danger')
        return redirect(url_for('company.dashboard'))
    
    equipment_requests = EquipmentRequest.query.filter_by(
        project_id=project_id
    ).order_by(EquipmentRequest.created_at.desc()).all()
    
    return render_template('projects/equipment_requests.html',
                         project=project,
                         equipment_requests=equipment_requests,
                         now=datetime.now())


@project_bp.route('/<int:project_id>/equipment-request/create', methods=['GET', 'POST'])
@login_required
def create_equipment_request(project_id):
    """إنشاء طلب معدات جديد"""
    from app.models import EquipmentRequest, EquipmentRequestItem, Resource
    
    project = Project.query.get_or_404(project_id)
    
    if project.org_id != current_user.org_id:
        flash('غير مصرح بالوصول', 'danger')
        return redirect(url_for('company.dashboard'))
    
    # جلب المعدات التي لم يتم طلبها بالكامل بعد
    project_equipment = []
    
    # جلب جميع المعدات المستخدمة في المشروع
    activities = Activity.query.filter_by(project_id=project_id).all()
    equipment_ids = set()
    
    for activity in activities:
        for assignment in activity.resources:
            if assignment.resource and assignment.resource.resource_type == 'equipment':
                equipment_ids.add(assignment.resource_id)
    
    # المعدات التي تم طلبها بالفعل
    existing_requests = EquipmentRequest.query.filter_by(project_id=project_id).all()
    requested_equipment_ids = set()
    
    for req in existing_requests:
        for item in req.items:
            if item.equipment_id:
                requested_equipment_ids.add(item.equipment_id)
    
    # المعدات المطلوب عرضها
    for equipment_id in equipment_ids:
        if equipment_id not in requested_equipment_ids:
            equipment = Resource.query.get(equipment_id)
            if equipment:
                project_equipment.append(equipment)
    
    # جلب جميع الموردين
    from app.models import User
    suppliers = User.query.filter_by(
        org_id=project.org_id,
        role='supplier',
        is_user_active=True
    ).all()
    
    today = datetime.now().date()
    default_date = today + timedelta(days=7)
    default_date_str = default_date.strftime('%Y-%m-%d')
    
    if request.method == 'POST':
        supplier_id = request.form.get('supplier_id')
        required_date = datetime.strptime(request.form.get('required_date'), '%Y-%m-%d').date()
        site_location = request.form.get('site_location')
        notes = request.form.get('notes')
        
        equipment_ids_selected = request.form.getlist('equipment_ids[]')
        quantities = request.form.getlist('quantities[]')
        
        if not equipment_ids_selected:
            flash('الرجاء اختيار معدة واحدة على الأقل', 'danger')
            return redirect(url_for('projects.create_equipment_request', project_id=project_id))
        
        try:
            equipment_request = EquipmentRequest(
                org_id=project.org_id,
                project_id=project_id,
                supplier_id=supplier_id,
                equipment_items=[],
                required_date=required_date,
                status='pending',
                site_location=site_location or (project.location.site_name if project.location else ''),
                notes=notes,
                created_by=current_user.id
            )
            db.session.add(equipment_request)
            db.session.flush()
            
            equipment_list = []
            for i in range(len(equipment_ids_selected)):
                equipment_id = int(equipment_ids_selected[i])
                quantity = float(quantities[i]) if quantities[i] else 0
                
                if quantity <= 0:
                    continue
                
                equipment = Resource.query.get(equipment_id)
                if equipment:
                    request_item = EquipmentRequestItem(
                        request_id=equipment_request.id,
                        equipment_id=equipment.id,
                        equipment_name=equipment.name,
                        equipment_type=equipment.equipment_type,
                        equipment_code=equipment.resource_id,
                        unit=equipment.unit,
                        required_quantity=quantity,
                        remaining_quantity=quantity
                    )
                    db.session.add(request_item)
                    
                    equipment_list.append({
                        'id': equipment.id,
                        'name': equipment.name,
                        'quantity': quantity,
                        'unit': equipment.unit
                    })
            
            equipment_request.equipment_items = equipment_list
            
            db.session.commit()
            
            from app.services.notification_service import NotificationService
            NotificationService.equipment_request_created(equipment_request)
            
            flash('تم إنشاء طلب المعدات بنجاح', 'success')
            return redirect(url_for('projects.project_equipment_requests', project_id=project_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ: {str(e)}', 'danger')
    
    return render_template('projects/create_equipment_request.html',
                         project=project,
                         equipment_list=project_equipment,
                         suppliers=suppliers,
                         default_date_str=default_date_str,
                         now=datetime.now())


@project_bp.route('/equipment-delivery/<int:delivery_id>/confirm-page', methods=['GET'])
@login_required
def confirm_equipment_delivery_page(delivery_id):
    """صفحة تأكيد تسليم المعدات"""
    from app.models import EquipmentDelivery, EquipmentRequest, EquipmentRequestItem
    
    delivery = EquipmentDelivery.query.get_or_404(delivery_id)
    equipment_request = delivery.request
    
    if equipment_request.project.org_id != current_user.org_id and current_user.role != 'org_admin':
        flash('غير مصرح', 'danger')
        return redirect(url_for('company.dashboard'))
    
    request_items = EquipmentRequestItem.query.filter_by(
        request_id=equipment_request.id
    ).all()
    
    return render_template('projects/confirm_equipment_delivery.html',
                         delivery=delivery,
                         equipment_request=equipment_request,
                         request_items=request_items,
                         project=equipment_request.project,
                         now=datetime.now())






