# app/routes/resource_routes.py

from flask import render_template, request, redirect, url_for, flash, jsonify, g,send_file
from flask_login import login_required, current_user
from app.models import db
from app.services.resource_service import ResourceService
from app.models import Resource, ActivityResource,Organization,Notification,Project,ResourceDelivery,Meeting,Issue,TaskPlanning,Task
from app.models import TaskResource,User,Activity
from app.routes import resource_bp 
import pandas as pd
import io
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json
from app.services.update_service import UpdateService
from datetime import datetime, timedelta,date
@resource_bp.before_request
def load_company():
    if current_user.is_authenticated:
        g.company = Organization.query.get(current_user.org_id)
        g.notifications_count = Notification.query.filter_by(
            user_id=current_user.id, 
            is_read=False
        ).count()
        g.delayed_tasks_count = Task.query.join(Project).filter(
            Task.status.in_(['pending', 'in_progress']),
            TaskPlanning.planned_finish < date.today()
        ).count()
        g.pending_deliveries_count = ResourceDelivery.query.filter_by(
            status='pending'
        ).count() if current_user.role in ['org_admin', 'project_manager'] else 0
        
        # إضافة إحصائيات الموارد
        g.low_stock_resources = Resource.query.filter(
            Resource.available_quantity < Resource.minimum_quantity
        ).count() if hasattr(Resource, 'minimum_quantity') else 0
        # ⭐ إحصائيات الاجتماعات القادمة
        # ========== الاجتماعات القادمة ==========
        today = datetime.now().date()
        next_week = today + timedelta(days=7)
        
        upcoming_meetings_query = Meeting.query.filter(
            Meeting.project.has(org_id=current_user.org_id),
            Meeting.scheduled_date >= today,
            Meeting.scheduled_date <= next_week,
            Meeting.status == 'scheduled'
        ).order_by(Meeting.scheduled_date)
        
        # الحصول على القائمة للعرض
        g.upcoming_meetings = upcoming_meetings_query.limit(5).all()
        
        # ✅ حساب العدد بشكل صحيح
        g.upcoming_meetings_count = upcoming_meetings_query.count()
        
        # ========== القضايا المفتوحة ==========
        open_issues_count = Issue.query.join(Project).filter(
            Project.org_id == current_user.org_id,
            Issue.status.in_(['open', 'in_progress'])
        ).count()
        g.open_issues_count = open_issues_count
        
        # آخر 5 قضايا
        recent_issues = Issue.query.join(Project).filter(
            Project.org_id == current_user.org_id
        ).order_by(Issue.reported_date.desc()).limit(5).all()
        g.recent_issues = recent_issues
        
        # إحصائيات القضايا
        g.issues_stats = {
            'open': Issue.query.join(Project).filter(
                Project.org_id == current_user.org_id, 
                Issue.status == 'open'
            ).count(),
            'in_progress': Issue.query.join(Project).filter(
                Project.org_id == current_user.org_id, 
                Issue.status == 'in_progress'
            ).count(),
            'total': Issue.query.join(Project).filter(
                Project.org_id == current_user.org_id
            ).count()
        }
    else:
        g.company = None
        g.upcoming_meetings =None
        g.recent_issues=None
        g.delayed_tasks_count = 0
        g.notifications_count = 0
        g.pending_deliveries_count=0
        g.low_stock_resources=0
        g.upcoming_meetings_count=0
        g.open_issues_count =0
        g.issues_stats={}

@resource_bp.route('/')
@login_required
def index():
    """صفحة إدارة الموارد الرئيسية"""
    service = ResourceService(current_user.org_id)
    result = service.get_resource_summary()
    
    # إضافة إحصائيات إضافية
    summary = result['summary']
    summary['total_allocated'] = sum(r.get('total_assigned', 0) for r in summary['resources_list'])
    summary['total_cost'] = sum(r.get('cost_per_unit', 0) * r.get('total_assigned', 0) for r in summary['resources_list'])
    
    return render_template('resources/index.html',
                         resources=summary['resources_list'],
                         summary=summary,
                         now=datetime.now())


@resource_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    """إنشاء مورد جديد"""
    if request.method == 'POST':
        data = {
            'resource_id': request.form.get('resource_id'),
            'name': request.form.get('name'),
            'resource_type': request.form.get('resource_type'),
            'unit': request.form.get('unit'),
            'cost_per_unit': float(request.form.get('cost_per_unit', 0)),
            'available_quantity': float(request.form.get('available_quantity', 0)),
            'currency': request.form.get('currency', 'SAR'),
            'calendar_id': request.form.get('calendar_id')
        }
        
        # حقول خاصة حسب النوع
        if data['resource_type'] == 'labor':
            data['employeeId'] = request.form.get('employeeId')
            data['specialization'] = request.form.get('specialization')
            data['skills'] = request.form.getlist('skills')
        elif data['resource_type'] == 'equipment':
            data['equipment_type'] = request.form.get('equipment_type')
            data['equipment_model'] = request.form.get('equipment_model')
            data['supplier_id'] = request.form.get('supplier_id')
        elif data['resource_type'] == 'material':
            data['material_type'] = request.form.get('material_type')
            data['supplier_id'] = request.form.get('supplier_id')
        
        service = ResourceService(current_user.org_id)
        result = service.create_resource(data)
        
        if result['success']:
            flash('تم إنشاء المورد بنجاح', 'success')
            return redirect(url_for('resource.index'))
        else:
            flash(result['error'], 'danger')
    
    # جلب الموردين والكالندرز للاختيار
    from app.models import Calendar, Supplier
    calendars = Calendar.query.filter_by(org_id=current_user.org_id).all()
    suppliers = User.query.filter_by(org_id=current_user.org_id,role="supplier").all()
    users = User.query.filter_by(org_id=current_user.org_id,role="employee").all()
    return render_template('resources/create.html',
                         calendars=calendars,
                         suppliers=suppliers,
                         users=users,
                         now=datetime.now())

@resource_bp.route('/createmultipal', methods=['GET', 'POST'])
@login_required
def createmultipal():
    """إنشاء موارد جديدة (يدعم موارد متعددة)"""
    if request.method == 'POST':
        resources_data = []
        
        # استخراج جميع الموارد من النموذج
        resource_keys = [key for key in request.form.keys() if key.startswith('resources[')]
        
        # استخراج الفهارس الفريدة
        indices = set()
        for key in resource_keys:
            import re
            match = re.search(r'resources\[(\d+)\]', key)
            if match:
                indices.add(int(match.group(1)))
        
        created_resources = []
        errors = []
        
        for idx in sorted(indices):
            try:
                # جمع بيانات المورد
                resource_data = {
                    'resource_id': request.form.get(f'resources[{idx}][resource_id]'),
                    'name': request.form.get(f'resources[{idx}][name]'),
                    'resource_type': request.form.get(f'resources[{idx}][resource_type]'),
                    'unit': request.form.get(f'resources[{idx}][unit]'),
                    'cost_per_unit': float(request.form.get(f'resources[{idx}][cost_per_unit]', 0)),
                    'available_quantity': float(request.form.get(f'resources[{idx}][available_quantity]', 0)),
                    'currency': request.form.get(f'resources[{idx}][currency]', 'SAR'),
                    'calendar_id': request.form.get(f'resources[{idx}][calendar_id]') or None,
                    'org_id': current_user.org_id,
                    'creator_id': current_user.id,
                    'is_active': True
                }
                
                # التحقق من البيانات الأساسية
                if not resource_data['resource_id']:
                    errors.append(f"المورد #{idx + 1}: كود المورد مطلوب")
                    continue
                    
                if not resource_data['name']:
                    errors.append(f"المورد #{idx + 1}: اسم المورد مطلوب")
                    continue
                    
                if not resource_data['resource_type']:
                    errors.append(f"المورد #{idx + 1}: نوع المورد مطلوب")
                    continue
                
                # التحقق من وجود الكود
                existing = Resource.query.filter_by(
                    org_id=current_user.org_id,
                    resource_id=resource_data['resource_id']
                ).first()
                
                if existing:
                    errors.append(f"المورد #{idx + 1}: كود {resource_data['resource_id']} موجود بالفعل")
                    continue
                
                # إنشاء مورد جديد
                resource = Resource(**resource_data)
                
                # إضافة الحقول الخاصة حسب النوع
                if resource_data['resource_type'] == 'labor':
                    resource.employee_id = request.form.get(f'resources[{idx}][employee_id]')
                    resource.specialization = request.form.get(f'resources[{idx}][specialization]')
                    skills = request.form.get(f'resources[{idx}][skills]')
                    if skills:
                        resource.skills = skills.split('\n')
                    resource.experience_years = float(request.form.get(f'resources[{idx}][experience_years]', 0))
                    
                elif resource_data['resource_type'] == 'equipment':
                    resource.equipment_type = request.form.get(f'resources[{idx}][equipment_type]')
                    resource.equipment_model = request.form.get(f'resources[{idx}][equipment_model]')
                    resource.equipment_serial = request.form.get(f'resources[{idx}][equipment_serial]')
                    resource.manufacturer = request.form.get(f'resources[{idx}][manufacturer]')
                    resource.manufacturing_year = request.form.get(f'resources[{idx}][manufacturing_year]')
                    resource.last_maintenance = request.form.get(f'resources[{idx}][last_maintenance]')
                    resource.next_maintenance = request.form.get(f'resources[{idx}][next_maintenance]')
                    resource.maintenance_cycle = int(request.form.get(f'resources[{idx}][maintenance_cycle]', 30))
                    resource.supplier_id = request.form.get(f'resources[{idx}][supplier_id]') or None
                    
                elif resource_data['resource_type'] == 'material':
                    resource.material_type = request.form.get(f'resources[{idx}][material_type]')
                    resource.material_grade = request.form.get(f'resources[{idx}][material_grade]')
                    resource.supplier_id = request.form.get(f'resources[{idx}][supplier_id]') or None
                    resource.minimum_quantity = float(request.form.get(f'resources[{idx}][minimum_quantity]', 0))
                    resource.maximum_quantity = float(request.form.get(f'resources[{idx}][maximum_quantity]', 0))
                    resource.reorder_quantity = float(request.form.get(f'resources[{idx}][reorder_quantity]', 0))
                
                db.session.add(resource)
                created_resources.append(resource_data['name'])
                
            except Exception as e:
                errors.append(f"المورد #{idx + 1}: {str(e)}")
        
        if created_resources:
            db.session.commit()
            flash(f'✅ تم إنشاء {len(created_resources)} مورد بنجاح: {", ".join(created_resources)}', 'success')
        
        if errors:
            flash(f'⚠️ الأخطاء:<br>{"<br>".join(errors)}', 'warning')
        
        if created_resources:
            return redirect(url_for('resource.index'))
        else:
            return redirect(url_for('resource.create'))
    
    # GET request - عرض النموذج
    from app.models import Calendar
    calendars = Calendar.query.filter_by(org_id=current_user.org_id, is_active=True).all()
    employees = User.query.filter_by(org_id=current_user.org_id, role='employee', is_user_active=True).all()
    suppliers = User.query.filter_by(org_id=current_user.org_id, role='supplier', is_user_active=True).all()
    
    return render_template('resources/create_multipal.html',
                         calendars=calendars,
                         employees=employees,
                         suppliers=suppliers,
                         now=datetime.now())


@resource_bp.route('/<int:resource_id>')
@login_required
def view(resource_id):
    """عرض تفاصيل مورد"""
    service = ResourceService(current_user.org_id)
    result = service.get_resource(resource_id)
    
    if not result['success']:
        flash(result['error'], 'danger')
        return redirect(url_for('resource.index'))
    
    resource = result['resource']
    
    # جلب التعيينات المرتبطة
    assignments = ActivityResource.query.filter_by(resource_id=resource_id).all()
    task_assignments = TaskResource.query.filter_by(resource_id=resource_id).all()
    
    return render_template('resources/view.html',
                         resource=resource,
                         assignments=assignments,
                         task_assignments=task_assignments,
                         now=datetime.now())


@resource_bp.route('/activity/<int:activity_id>/resources')
@login_required
def activity_resources(activity_id):
    """عرض موارد النشاط"""
    activity = Activity.query.get_or_404(activity_id)
    
    # جلب موارد النشاط
    activity_resources = []
    for ar in activity.resources:
        activity_resources.append({
            'id': ar.id,
            'resource_id': ar.resource.id,
            'resource_name': ar.resource.name,
            'resource_code': ar.resource.resource_id,
            'resource_type': ar.resource.resource_type,
            'unit': ar.resource.unit,
            'planned_quantity': ar.planned_quantity,
            'actual_quantity': ar.actual_quantity,
            'remaining_quantity': ar.remaining_quantity,
            'cost_per_unit': ar.resource.cost_per_unit,
            'planned_cost': ar.planned_cost,
            'actual_cost': ar.actual_cost,
            'utilization': (ar.actual_quantity / ar.planned_quantity * 100) if ar.planned_quantity > 0 else 0
        })
    
    # جلب جميع الموارد المتاحة للاختيار
    all_resources = Resource.query.filter_by(
        org_id=current_user.org_id,
        is_active=True
    ).all()
    
    return render_template('resources/activity_resources.html',
                         activity=activity,
                         resources=activity_resources,
                         all_resources=all_resources,
                         now=datetime.now())
@resource_bp.route('/api/activity/<int:activity_id>/resources')
@login_required
def api_activity_resources(activity_id):
    """API لجلب موارد النشاط"""
    activity = Activity.query.get_or_404(activity_id)
    
    resources_data = []
    for ar in activity.resources:
        resources_data.append({
            'id': ar.id,
            'resource_id': ar.resource.id,
            'resource_name': ar.resource.name,
            'resource_code': ar.resource.resource_id,
            'resource_type': ar.resource.resource_type,
            'unit': ar.resource.unit,
            'planned_quantity': ar.planned_quantity,
            'actual_quantity': ar.actual_quantity,
            'allocated_quantity': ar.allocated_quantity,
            'remaining_quantity': ar.remaining_quantity,
            'cost_per_unit': ar.resource.cost_per_unit,
            'planned_cost': ar.planned_cost,
            'actual_cost': ar.actual_cost
        })
    
    return jsonify({
        'success': True,
        'resources': resources_data
    })


@resource_bp.route('/api/activity/<int:activity_id>/assign', methods=['POST'])
@login_required
def api_assign_resource_to_activity(activity_id):
    """API لتخصيص مورد لنشاط"""
    data = request.get_json()
    resource_id = data.get('resource_id')
    quantity = data.get('quantity', 0)
    
    if not resource_id:
        return jsonify({'success': False, 'error': 'المورد مطلوب'}), 400
    
    if quantity <= 0:
        return jsonify({'success': False, 'error': 'الكمية يجب أن تكون أكبر من صفر'}), 400
    
    resource = Resource.query.get_or_404(resource_id)
    activity = Activity.query.get_or_404(activity_id)
    
    # التحقق من الكمية المتاحة
    if resource.available_quantity < quantity:
        return jsonify({
            'success': False,
            'error': f'الكمية غير متوفرة. المتاح: {resource.available_quantity} {resource.unit}'
        }), 400
    
    try:
        # البحث عن تخصيص موجود
        existing = ActivityResource.query.filter_by(
            activity_id=activity_id,
            resource_id=resource_id
        ).first()
        
        if existing:
            # تحديث الكمية
            existing.planned_quantity += quantity
            existing.planned_cost = existing.planned_quantity * resource.cost_per_unit
            existing.remaining_quantity = existing.planned_quantity - existing.actual_quantity
        else:
            # إنشاء تخصيص جديد
            assignment = ActivityResource(
                activity_id=activity_id,
                resource_id=resource_id,
                planned_quantity=quantity,
                planned_cost=quantity * resource.cost_per_unit,
                remaining_quantity=quantity,
                allocated_quantity=quantity,
                created_by=current_user.id
            )
            db.session.add(assignment)
        
        # تحديث الكمية المخصصة في المورد
        resource.total_allocated += quantity
        resource.update_utilization()
        
        db.session.commit()
        # ✅ تحديث المؤشرات
        UpdateService.update_activity_metrics(activity)
        UpdateService.update_project_metrics(activity.project)
        UpdateService.update_resource_metrics(resource.id)
        return jsonify({
            'success': True,
            'message': 'تم تخصيص المورد بنجاح',
            'remaining_quantity': resource.available_quantity - resource.total_allocated
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@resource_bp.route('/api/activity-resource/<int:assignment_id>/update', methods=['POST'])
@login_required
def api_update_activity_resource(assignment_id):
    """API لتحديث كمية مورد في نشاط"""
    data = request.get_json()
    quantity = data.get('quantity', 0)
    
    if quantity <= 0:
        return jsonify({'success': False, 'error': 'الكمية يجب أن تكون أكبر من صفر'}), 400
    
    assignment = ActivityResource.query.get_or_404(assignment_id)
    activity = assignment.activity
    # الفرق في الكمية
    diff = quantity - assignment.planned_quantity
    
    if diff > 0:
        # زيادة الكمية
        resource = assignment.resource
        if resource.available_quantity - resource.total_allocated < diff:
            return jsonify({
                'success': False,
                'error': f'الكمية غير متوفرة. المتاح: {resource.available_quantity - resource.total_allocated}'
            }), 400
        
        resource.total_allocated += diff
    
    elif diff < 0:
        # تقليل الكمية
        assignment.resource.total_allocated += diff  # diff سالب
    
    assignment.planned_quantity = quantity
    assignment.planned_cost = quantity * assignment.resource.cost_per_unit
    assignment.remaining_quantity = quantity - assignment.actual_quantity
    
    db.session.commit()
    # ✅ تحديث المؤشرات
    UpdateService.update_activity_metrics(activity)
    UpdateService.update_project_metrics(activity.project)
    UpdateService.update_resource_metrics(assignment.resource_id)
    
    return jsonify({'success': True})


@resource_bp.route('/api/activity-resource/<int:assignment_id>/remove', methods=['POST'])
@login_required
def api_remove_activity_resource(assignment_id):
    """API لإزالة مورد من نشاط"""
    assignment = ActivityResource.query.get_or_404(assignment_id)
    
    # تحرير الكمية المخصصة
    assignment.resource.total_allocated -= assignment.planned_quantity
    assignment.resource.update_utilization()
    
    db.session.delete(assignment)
    db.session.commit()
    
    return jsonify({'success': True})

@resource_bp.route('/api/activity/<int:activity_id>/allocate', methods=['POST'])
@login_required
def api_allocate_to_activity(activity_id):
    """API لتخصيص مورد لنشاط"""
    data = request.get_json()
    resource_id = data.get('resource_id')
    quantity = data.get('quantity', 0)
    
    service = ResourceService(current_user.org_id)
    result = service.allocate_resource_to_activity(activity_id, resource_id, quantity)
    
    return jsonify(result)


@resource_bp.route('/api/task/<int:task_id>/allocate', methods=['POST'])
@login_required
def api_allocate_to_task(task_id):
    """API لتوزيع مورد من النشاط على مهمة"""
    data = request.get_json()
    activity_resource_id = data.get('activity_resource_id')
    quantity = data.get('quantity', 0)
    allocation_percentage = data.get('allocation_percentage', 0)
    
    service = ResourceService(current_user.org_id)
    result = service.allocate_resource_to_task(
        task_id, activity_resource_id, quantity, allocation_percentage
    )
    
    return jsonify(result)


@resource_bp.route('/api/task/<int:task_id>/resources')
@login_required
def api_task_resources(task_id):
    """API لجلب موارد المهمة"""
    service = ResourceService(current_user.org_id)
    result = service.get_task_resources(task_id)
    
    return jsonify(result)
# أضف هذه الروالت إلى resource_routes.py

# app/routes/resource_routes.py - إضافة راوت التعديل

@resource_bp.route('/<int:resource_id>/edit', methods=['GET', 'POST'])
@login_required
def edit(resource_id):
    """تعديل مورد"""
    service = ResourceService(current_user.org_id)
    result = service.get_resource(resource_id)
    
    if not result['success']:
        flash(result['error'], 'danger')
        return redirect(url_for('resource.index'))
    
    resource = result['resource']
    
    # جلب التقويمات والموردين للاختيار
    from app.models.primavera_models import Calendar
    from app.models.project_models import Supplier
    
    calendars = Calendar.query.filter_by(org_id=current_user.org_id).all()
    suppliers= User.query.filter_by(org_id=current_user.org_id,role="supplier").all()
    users= User.query.filter_by(org_id=current_user.org_id,role="employee").all()
    
    if request.method == 'POST':
        data = {
            'name': request.form.get('name'),
            'resource_type': request.form.get('resource_type'),
            'unit': request.form.get('unit'),
            'cost_per_unit': float(request.form.get('cost_per_unit', 0)),
            'available_quantity': float(request.form.get('available_quantity', 0)),
            'currency': request.form.get('currency', 'SAR'),
            'calendar_id': request.form.get('calendar_id') or None,
            'is_active': request.form.get('is_active') == 'on'
        }
        
        # حقول خاصة حسب النوع
        if data['resource_type'] == 'labor':
            data['employeeId'] = request.form.get('employeeId')
            data['specialization'] = request.form.get('specialization')
            data['specialization'] = request.form.get('specialization')
            data['skills'] = request.form.get('skills', '').split(',') if request.form.get('skills') else []
            data['experience_years'] = float(request.form.get('experience_years', 0))
        elif data['resource_type'] == 'equipment':
            data['equipment_type'] = request.form.get('equipment_type')
            data['equipment_model'] = request.form.get('equipment_model')
            data['manufacturer'] = request.form.get('manufacturer')
        elif data['resource_type'] == 'material':
            data['material_type'] = request.form.get('material_type')
            data['material_grade'] = request.form.get('material_grade')
            data['supplier_id'] = request.form.get('supplier_id') or None
            data['minimum_quantity'] = float(request.form.get('minimum_quantity', 0))
            data['reorder_quantity'] = float(request.form.get('reorder_quantity', 0))
        
        result = service.update_resource(resource_id, data)
        
        if result['success']:
            flash('تم تحديث المورد بنجاح', 'success')
            return redirect(url_for('resource.view', resource_id=resource_id))
        else:
            flash(result['error'], 'danger')
    
    return render_template('resources/edit.html',
                         resource=resource,
                         calendars=calendars,
                         suppliers=suppliers,
                         users=users,
                         now=datetime.now())


@resource_bp.route('/api/<int:resource_id>/delete', methods=['POST'])
@login_required
def delete(resource_id):
    """حذف مورد"""
    service = ResourceService(current_user.org_id)
    result = service.delete_resource(resource_id)
    
    return jsonify(result)

# ============================================
# طرق إضافة موارد متعددة
# ============================================

@resource_bp.route('/bulk-import')
@login_required
def bulk_import():
    """صفحة استيراد موارد متعددة من Excel"""
    return render_template('resources/bulk_import.html', now=datetime.now())


@resource_bp.route('/bulk-add')
@login_required
def bulk_add():
    """صفحة إضافة موارد متعددة يدوياً"""
    suppliers = User.query.filter_by(org_id=current_user.org_id,role="supplier").all()
    return render_template('resources/bulk_add.html',suppliers=suppliers, now=datetime.now())


@resource_bp.route('/api/parse-excel', methods=['POST'])
@login_required
def parse_excel():
    """معالجة ملف Excel واستخراج البيانات"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'لم يتم رفع ملف'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'الملف فارغ'}), 400
    
    try:
        # قراءة الملف
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        # تحويل البيانات إلى قواميس
        data = []
        for _, row in df.iterrows():
            # تحديد نوع المورد
            resource_type = row.get('النوع') or row.get('type')
            if resource_type:
                resource_type = map_resource_type(str(resource_type))
            else:
                resource_type = 'non_labor'
            
            resource = {
                'code': str(row.get('الكود') or row.get('code', '')),
                'name': str(row.get('الاسم') or row.get('name', '')),
                'type': resource_type,
                'unit': str(row.get('الوحدة') or row.get('unit', '')),
                'cost': float(row.get('التكلفة') or row.get('cost', 0)),
                'quantity': float(row.get('الكمية') or row.get('quantity', 0)),
                'description': str(row.get('الوصف') or row.get('description', '')),
                'specialization': str(row.get('التخصص') or row.get('specialization', '')),
                'equipment_type': str(row.get('نوع_المعدة') or row.get('equipment_type', '')),
                'material_type': str(row.get('نوع_المادة') or row.get('material_type', ''))
            }
            
            # تنظيف البيانات الفارغة
            for key in resource:
                if resource[key] in ['', 'nan', 'None', 'NaN']:
                    resource[key] = None
            
            data.append(resource)
        
        return jsonify({'success': True, 'data': data, 'count': len(data)})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@resource_bp.route('/api/bulk-create', methods=['POST'])
@login_required
def bulk_create_resource():
    """إنشاء مورد واحد من البيانات المستوردة"""
    data = request.get_json()
    
    try:
        # التحقق من البيانات الأساسية
        if not data.get('name'):
            return jsonify({'success': False, 'error': 'اسم المورد مطلوب'}), 400
        
        # التحقق من عدم تكرار الكود
        if data.get('code'):
            existing = Resource.query.filter_by(
                org_id=current_user.org_id,
                resource_id=data['code']
            ).first()
            
            if existing:
                return jsonify({'success': False, 'error': 'كود المورد موجود مسبقاً'}), 400
        else:
            # إنشاء كود تلقائي
            data['code'] = f"RES-{uuid.uuid4().hex[:8].upper()}"
        
        # إنشاء المورد
        resource = Resource(
            org_id=current_user.org_id,
            resource_id=data['code'],
            name=data['name'],
            description=data.get('description'),
            resource_type=data.get('type', 'non_labor'),
            unit=data.get('unit', 'piece'),
            cost_per_unit=data.get('cost', 0),
            available_quantity=data.get('quantity', 0),
            is_active=True,
            creator_id=current_user.id
        )
        
        # إضافة حقول خاصة حسب النوع
        if data.get('type') == 'labor':
            resource.specialization = data.get('specialization')
        elif data.get('type') == 'equipment':
            resource.equipment_type = data.get('equipment_type')
        elif data.get('type') == 'material':
            resource.material_type = data.get('material_type')
        
        db.session.add(resource)
        db.session.commit()
        
        return jsonify({'success': True, 'id': resource.id})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@resource_bp.route('/api/bulk-create-multiple', methods=['POST'])
@login_required
def bulk_create_multiple():
    """إنشاء موارد متعددة دفعة واحدة مع دعم الحقول الخاصة"""
    data = request.get_json()
    resources_data = data.get('resources', [])
    
    if not resources_data:
        return jsonify({'success': False, 'error': 'لا توجد بيانات'}), 400
    
    results = {
        'success': [],
        'failed': []
    }
    
    for resource_data in resources_data:
        try:
            # التحقق من البيانات الأساسية
            if not resource_data.get('name'):
                results['failed'].append({
                    'data': resource_data,
                    'error': 'اسم المورد مطلوب'
                })
                continue
            
            # التحقق من عدم تكرار الكود
            if resource_data.get('code'):
                existing = Resource.query.filter_by(
                    org_id=current_user.org_id,
                    resource_id=resource_data['code']
                ).first()
                
                if existing:
                    results['failed'].append({
                        'data': resource_data,
                        'error': 'كود المورد موجود مسبقاً'
                    })
                    continue
            else:
                # إنشاء كود تلقائي
                resource_data['code'] = f"RES-{uuid.uuid4().hex[:8].upper()}"
            
            # إنشاء المورد
            resource = Resource(
                org_id=current_user.org_id,
                resource_id=resource_data['code'],
                name=resource_data['name'],
                resource_type=resource_data.get('type', 'non_labor'),
                unit=resource_data.get('unit', 'piece'),
                cost_per_unit=resource_data.get('cost_per_unit', 0),
                available_quantity=resource_data.get('available_quantity', 0),
                currency=resource_data.get('currency', 'SAR'),
                minimum_quantity=resource_data.get('minimum_quantity', 0),
                is_active=True,
                created_by=current_user.id
            )
            
            # إضافة حقول خاصة حسب النوع
            if resource_data.get('type') == 'labor':
                resource.specialization = resource_data.get('specialization')
                resource.skills = resource_data.get('skills', []) if isinstance(resource_data.get('skills'), list) else resource_data.get('skills', '').split(',')
                resource.experience_years = resource_data.get('experience_years', 0)
                
            elif resource_data.get('type') == 'equipment':
                resource.equipment_type = resource_data.get('equipment_type')
                resource.equipment_model = resource_data.get('equipment_model')
                resource.manufacturer = resource_data.get('manufacturer')
                
            elif resource_data.get('type') == 'material':
                resource.material_type = resource_data.get('material_type')
                resource.material_grade = resource_data.get('material_grade')
                resource.supplier_id = resource_data.get('supplier_id')
                resource.reorder_quantity = resource_data.get('reorder_quantity', 0)
            
            db.session.add(resource)
            db.session.flush()
            
            results['success'].append({
                'id': resource.id,
                'code': resource.resource_id,
                'name': resource.name,
                'type': resource.resource_type
            })
            
        except Exception as e:
            results['failed'].append({
                'data': resource_data,
                'error': str(e)
            })
            db.session.rollback()
    
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    
    return jsonify({
        'success': True,
        'results': results,
        'summary': {
            'total': len(resources_data),
            'success': len(results['success']),
            'failed': len(results['failed'])
        }
    })

@resource_bp.route('/download-template')
@login_required
def download_template():
    """تحميل قالب Excel للاستيراد"""
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "الموارد"
    
    # تحديد الأعمدة
    headers = [
        ('الكود', 'code'),
        ('الاسم', 'name'),
        ('النوع', 'type'),
        ('الوحدة', 'unit'),
        ('التكلفة', 'cost'),
        ('الكمية', 'quantity'),
        ('الوصف', 'description'),
        ('التخصص', 'specialization'),
        ('نوع_المعدة', 'equipment_type'),
        ('نوع_المادة', 'material_type')
    ]
    
    # تنسيق الرأس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
    
    for col, (header_ar, header_en) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header_ar)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
        # إضافة تعليق بالانجليزي
        ws.cell(row=2, column=col, value=f"({header_en})")
    
    # إضافة أمثلة
    examples = [
        ['RES-001', 'أحمد محمد', 'أحمد محمد', 'labor', 'hour', 50, 160, 'عامل بناء', 'بناء', '', ''],
        ['RES-002', 'حفار كوماتسو', 'حفار كوماتسو', 'equipment', 'day', 800, 1, 'حفار 200', '', 'حفار', ''],
        ['RES-003', 'حديد تسليح', 'حديد تسليح', 'material', 'ton', 2500, 50, 'حديد قطر 12', '', '', 'حديد']
    ]
    
    for row, example in enumerate(examples, 3):
        for col, value in enumerate(example, 1):
            ws.cell(row=row, column=col, value=value)
    
    # ضبط عرض الأعمدة
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # حفظ في الذاكرة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='resources_template.xlsx'
    )


def map_resource_type(resource_type):
    """تحويل نوع المورد من عربي إلى إنجليزي"""
    type_map = {
        'عمال': 'labor',
        'عمالة': 'labor',
        'موارد بشرية': 'labor',
        'بشري': 'labor',
        'معدات': 'equipment',
        'آلات': 'equipment',
        'مواد': 'material',
        'خامات': 'material',
        'ادوات': 'material',
        'أدوات': 'material',
        'labor': 'labor',
        'equipment': 'equipment',
        'material': 'material',
        'non_labor': 'non_labor'
    }
    return type_map.get(str(resource_type).lower(), 'non_labor')

@resource_bp.route('/project/<int:project_id>/resources')
@login_required
def project_resources(project_id):
    """عرض موارد المشروع"""
    project = Project.query.get_or_404(project_id)
    
    # جلب جميع الموارد المرتبطة بالمشروع
    project_resources = []
    
    # الموارد المرتبطة عبر الأنشطة
    activities = Activity.query.filter_by(project_id=project_id).all()
    for activity in activities:
        for ar in activity.resources:
            project_resources.append({
                'id': ar.id,
                'resource_id': ar.resource.id,
                'resource_name': ar.resource.name,
                'resource_type': ar.resource.resource_type,
                'unit': ar.resource.unit,
                'planned_quantity': ar.planned_quantity,
                'actual_quantity': ar.actual_quantity,
                'allocated_quantity': ar.allocated_quantity,
                'cost_per_unit': ar.resource.cost_per_unit,
                'total_cost': ar.planned_cost,
                'activity_id': activity.id,
                'activity_name': activity.activity_name
            })
    
    # الموارد المتاحة للإضافة
    available_resources = Resource.query.filter_by(
        org_id=current_user.org_id,
        is_active=True
    ).all()
    
    return render_template('resources/project_resources.html',
                         project=project,
                         resources=project_resources,
                         available_resources=available_resources,
                         now=datetime.now())
@resource_bp.route('/api/project/<int:project_id>/resources')
@login_required
def api_project_resources(project_id):
    """API لجلب موارد المشروع"""
    project = Project.query.get_or_404(project_id)
    
    # جلب الموارد عبر الأنشطة
    resources_data = []
    activities = Activity.query.filter_by(project_id=project_id).all()
    
    for activity in activities:
        for ar in activity.resources:
            resources_data.append({
                'assignment_id': ar.id,
                'resource_id': ar.resource.id,
                'resource_name': ar.resource.name,
                'resource_code': ar.resource.resource_id,
                'resource_type': ar.resource.resource_type,
                'unit': ar.resource.unit,
                'planned_quantity': ar.planned_quantity,
                'actual_quantity': ar.actual_quantity,
                'allocated_quantity': ar.allocated_quantity,
                'cost_per_unit': ar.resource.cost_per_unit,
                'planned_cost': ar.planned_cost,
                'actual_cost': ar.actual_cost,
                'activity_id': activity.id,
                'activity_name': activity.activity_name,
                'status': 'allocated'
            })
    
    return jsonify({
        'success': True,
        'resources': resources_data,
        'total': len(resources_data)
    })


@resource_bp.route('/api/project/<int:project_id>/assign', methods=['POST'])
@login_required
def api_assign_resource_to_project(project_id):
    """API لتخصيص مورد لمشروع مباشرة"""
    data = request.get_json()
    resource_id = data.get('resource_id')
    quantity = data.get('quantity', 0)
    
    if not resource_id:
        return jsonify({'success': False, 'error': 'المورد مطلوب'}), 400
    
    if quantity <= 0:
        return jsonify({'success': False, 'error': 'الكمية يجب أن تكون أكبر من صفر'}), 400
    
    resource = Resource.query.get_or_404(resource_id)
    project = Project.query.get_or_404(project_id)
    
    # التحقق من الكمية المتاحة
    if resource.available_quantity - resource.total_allocated < quantity:
        return jsonify({
            'success': False,
            'error': f'الكمية غير متوفرة. المتاح: {resource.available_quantity - resource.total_allocated} {resource.unit}'
        }), 400
    
    try:
        # إنشاء نشاط افتراضي للمشروع (اختياري)
        # أو يمكن تخزين المورد مباشرة في المشروع
        
        # تحديث المورد
        resource.total_allocated += quantity
        resource.update_utilization()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم تخصيص المورد للمشروع بنجاح'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@resource_bp.route('/api/project/<int:project_id>/bulk-import-preview', methods=['POST'])
@login_required
def api_bulk_import_preview(project_id):
    """معاينة بيانات الاستيراد للمشروع"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'لم يتم رفع ملف'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'الملف فارغ'}), 400
    
    try:
        # قراءة الملف
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        # معاينة البيانات
        preview_data = []
        for _, row in df.iterrows():
            # البحث عن المورد
            resource = None
            code = row.get('الكود') or row.get('code')
            if code:
                resource = Resource.query.filter_by(
                    org_id=current_user.org_id,
                    resource_id=code
                ).first()
            
            name = row.get('الاسم') or row.get('name')
            if not resource and name:
                resource = Resource.query.filter_by(
                    org_id=current_user.org_id,
                    name=name
                ).first()
            
            activity = None
            activity_name = row.get('النشاط') or row.get('activity')
            if activity_name:
                activity = Activity.query.filter_by(
                    project_id=project_id,
                    activity_name=activity_name
                ).first()
            
            preview_data.append({
                'code': code,
                'name': name,
                'activity': activity_name,
                'quantity': float(row.get('الكمية') or row.get('quantity', 0)),
                'unit': row.get('الوحدة') or row.get('unit', ''),
                'status': 'valid' if resource else 'invalid'
            })
        
        return jsonify({'success': True, 'data': preview_data})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@resource_bp.route('/api/project/<int:project_id>/bulk-import', methods=['POST'])
@login_required
def api_bulk_import_project_resources(project_id):
    """استيراد موارد متعددة للمشروع"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'لم يتم رفع ملف'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'الملف فارغ'}), 400
    
    project = Project.query.get_or_404(project_id)
    
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        results = {'success': [], 'failed': []}
        
        for _, row in df.iterrows():
            try:
                # البحث عن المورد
                code = row.get('الكود') or row.get('code')
                if code:
                    resource = Resource.query.filter_by(
                        org_id=current_user.org_id,
                        resource_id=code
                    ).first()
                else:
                    name = row.get('الاسم') or row.get('name')
                    resource = Resource.query.filter_by(
                        org_id=current_user.org_id,
                        name=name
                    ).first()
                
                if not resource:
                    results['failed'].append({
                        'name': name or code,
                        'error': 'المورد غير موجود'
                    })
                    continue
                
                quantity = float(row.get('الكمية') or row.get('quantity', 0))
                
                # البحث عن نشاط
                activity = None
                activity_name = row.get('النشاط') or row.get('activity')
                if activity_name:
                    activity = Activity.query.filter_by(
                        project_id=project_id,
                        activity_name=activity_name
                    ).first()
                
                # إنشاء تخصيص
                assignment = ActivityResource(
                    activity_id=activity.id if activity else None,
                    resource_id=resource.id,
                    planned_quantity=quantity,
                    planned_cost=quantity * resource.cost_per_unit,
                    remaining_quantity=quantity,
                    allocated_quantity=quantity,
                    created_by=current_user.id
                )
                db.session.add(assignment)
                
                resource.total_allocated += quantity
                resource.update_utilization()
                
                results['success'].append({
                    'name': resource.name,
                    'quantity': quantity,
                    'activity': activity.activity_name if activity else 'المشروع'
                })
                
            except Exception as e:
                results['failed'].append({
                    'name': row.get('الاسم') or row.get('name', 'غير معروف'),
                    'error': str(e)
                })
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'results': results,
            'summary': {
                'success': len(results['success']),
                'failed': len(results['failed'])
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@resource_bp.route('/project/<int:project_id>/import-template')
@login_required
def download_project_import_template(project_id):
    """تحميل قالب استيراد موارد للمشروع"""
    project = Project.query.get_or_404(project_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "موارد المشروع"
    
    headers = ['الكود', 'الاسم', 'النشاط', 'الكمية', 'الوحدة', 'ملاحظات']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    examples = [
        ['RES-001', 'حفار كوماتسو', 'حفر الأساسات', 2, 'يوم', 'مطلوب للتسليم'],
        ['RES-002', 'حديد تسليح', 'صب الخرسانة', 50, 'طن', ''],
        ['', 'عامل بناء', 'أعمال التشطيب', 10, 'عامل', '']
    ]
    
    for row, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            ws.cell(row=row, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'project_{project.project_code}_resources_template.xlsx'
    )